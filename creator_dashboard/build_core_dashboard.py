from __future__ import annotations

import csv
import json
import os
import re
import subprocess
from datetime import datetime, timedelta
from pathlib import Path
from urllib.parse import quote

from openpyxl import load_workbook

try:
    import tomllib  # type: ignore[attr-defined]
except ModuleNotFoundError:  # pragma: no cover
    tomllib = None


BASE_DIR = Path(__file__).resolve().parent
WORKBOOK_PATH = Path("/Users/apple/Desktop/达人多次合作监控看板_同步版.xlsx")
OUTPUT_JSON = BASE_DIR / "data" / "core_creator_dashboard.json"
OUTPUT_CSV = BASE_DIR / "data" / "core_creator_overview.csv"
OUTPUT_JS = BASE_DIR / "data" / "core_creator_dashboard.js"
CREATOR_POOL_JSON = BASE_DIR / "data" / "creator_pool.json"
PIPELINE_STATE_PATH = Path("/Users/apple/Documents/Playground/tiktok_shop_sync/data/pipeline_state.json")
AUTOMATION_TOML_PATH = Path(os.path.expanduser("~/.codex/automations/tiktok/automation.toml"))
LEVEL_SHEET_PATH = Path("/Users/apple/Downloads/达人等级底表/达人等级底表.xlsx")
SUPABASE_URL = "https://sbznfjnsirajqkkcwayj.supabase.co"
SUPABASE_ANON_KEY = "sb_publishable_tM67K7Mi1qDUkemhgzDuGg_dsdwitBT"
REST_PAGE_SIZE = 1000
SUPABASE_READ_TIMEOUT_SECONDS = float(os.environ.get("TIKTOK_PUBLIC_READ_TIMEOUT_SECONDS", "5"))
ANSI_ESCAPE_RE = re.compile(r"\x1B\[[0-?]*[ -/]*[@-~]")
CONTROL_CHAR_RE = re.compile(r"[\x00-\x08\x0B-\x0C\x0E-\x1F]")

OVERVIEW_HEADERS = [
    "达人ID",
    "达人名称",
    "平台",
    "粉丝量",
    "达人分层(L0/L1/L2/L3)",
    "达人类型",
    "品牌标签",
    "历史总GMV",
    "90天GMV",
    "近30天发布视频gmv",
    "最近合作日期",
    "当前合作状态",
    "近30天合作次数",
    "平均间隔天数",
    "距离上次发布天数",
    "近30天是否出单(Y/N)",
    "是否进入复投(Y/N)",
    "是否超时未合作",
    "平均交付时间",
    "优先级",
    "下一步动作",
    "负责人",
    "截止日期",
    "备注",
]

FOCUS_HEADERS = OVERVIEW_HEADERS.copy()

RECORD_HEADERS = [
    "达人ID",
    "达人名称",
    "合作日期",
    "产品",
    "视频链接",
    "内容类型",
    "是否出单(Y/N)",
    "订单数",
    "GMV",
    "佣金",
    "是否复投(Y/N)",
    "备注",
]


BRAND_ALIAS_MAP = {
    "letme home living": "LetMe",
    "letme": "LetMe",
    "spar.co jewelry": "Sparco",
    "sparco": "Sparco",
    "stypro.id": "Stypro",
    "stypro": "Stypro",
    "icyee indonesia": "ICYEE",
    "icyee": "ICYEE",
}

CREATOR_RAW_COLUMNS = [
    "stat_date",
    "store_tag",
    "range_label",
    "creator_name",
    "creator_handle",
    "creator_key",
    "affiliate_followers",
    "affiliate_gmv",
    "attributed_orders",
    "videos",
]

VIDEO_RAW_COLUMNS = [
    "stat_date",
    "store_tag",
    "creator_key",
    "post_date",
    "affiliate_video_gmv",
    "video_orders",
    "video_id",
    "video_link",
    "product_id",
    "product_name",
]


def normalize_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    text = str(value)
    text = ANSI_ESCAPE_RE.sub("", text)
    text = CONTROL_CHAR_RE.sub("", text)
    return text.strip()


def normalize_number(value: object) -> float:
    text = normalize_text(value)
    if not text:
        return 0.0
    text = text.replace("Rp", "").replace("rp", "").replace(" ", "")
    if "," in text and "." in text:
        text = text.replace(".", "").replace(",", ".")
    else:
        text = text.replace(",", "")
        if text.count(".") > 1:
            text = text.replace(".", "")
    try:
        return float(text)
    except ValueError:
        return 0.0


def normalize_int_string(value: object) -> str:
    num = normalize_number(value)
    if not num:
        return ""
    if abs(num - round(num)) < 1e-9:
        return str(int(round(num)))
    return normalize_text(value)


def normalize_handle(value: object) -> str:
    text = normalize_text(value).lower().lstrip("@")
    return re.sub(r"[^a-z0-9._]+", "", text)


def parse_date(value: object) -> datetime | None:
    text = normalize_text(value)
    if not text:
        return None
    for fmt in ("%Y-%m-%d", "%m/%d/%Y %H:%M", "%m/%d/%Y", "%Y/%m/%d"):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue
    return None


def parse_range_span_days(value: object) -> int:
    text = normalize_text(value)
    if not text:
        return 0
    if text.endswith("单日"):
        return 1
    match = re.search(r"(\d{4}-\d{2}-\d{2})至(\d{4}-\d{2}-\d{2})", text)
    if not match:
        return 0
    start = parse_date(match.group(1))
    end = parse_date(match.group(2))
    if not start or not end:
        return 0
    return max((end - start).days + 1, 0)


def sheet_headers(worksheet) -> list[str]:
    return [normalize_text(cell.value) for cell in worksheet[1]]


def rows_as_dicts(worksheet, start_row: int = 2) -> list[dict[str, object]]:
    headers = sheet_headers(worksheet)
    rows: list[dict[str, object]] = []
    for row in worksheet.iter_rows(min_row=start_row, values_only=True):
        record = {headers[index]: row[index] for index in range(min(len(headers), len(row)))}
        if not any(normalize_text(value) for value in record.values()):
            continue
        rows.append(record)
    return rows


def split_multi_value(value: object) -> list[str]:
    text = normalize_text(value)
    if not text:
        return []
    for delimiter in [" / ", "/", "，", ",", "；", ";"]:
        text = text.replace(delimiter, "|")
    return [item.strip() for item in text.split("|") if item.strip()]


def canonical_brand(value: object) -> str:
    text = normalize_text(value).lower()
    for needle, brand in BRAND_ALIAS_MAP.items():
        if needle in text:
            return brand
    return ""


def supabase_request_json(path: str, headers: dict[str, str] | None = None) -> tuple[object, dict[str, str]]:
    command = [
        "curl",
        "-sS",
        "--max-time",
        str(int(max(SUPABASE_READ_TIMEOUT_SECONDS, 1))),
        f"{SUPABASE_URL}/rest/v1{path}",
        "-H",
        f"apikey: {SUPABASE_ANON_KEY}",
        "-H",
        f"Authorization: Bearer {SUPABASE_ANON_KEY}",
    ]
    for key, value in (headers or {}).items():
        command.extend(["-H", f"{key}: {value}"])
    try:
        payload_text = subprocess.check_output(command, text=True).strip()
    except subprocess.CalledProcessError:
        raise RuntimeError("supabase_request_failed")
    if not payload_text:
        return [], {}
    return json.loads(payload_text), {}


def fetch_supabase_rows(table: str, columns: list[str], page_size: int = REST_PAGE_SIZE, extra_query: str = "") -> list[dict[str, object]]:
    encoded_select = quote(",".join(columns), safe=",")
    query_suffix = f"&{extra_query}" if extra_query else ""
    offset = 0
    rows: list[dict[str, object]] = []
    while True:
        try:
            batch, _headers = supabase_request_json(
                f"/{table}?select={encoded_select}{query_suffix}",
                headers={
                    "Range-Unit": "items",
                    "Range": f"{offset}-{offset + page_size - 1}",
                },
            )
        except (RuntimeError, TimeoutError, OSError, json.JSONDecodeError):
            return rows
        if not isinstance(batch, list):
            raise RuntimeError(f"Unexpected Supabase response for {table}")
        rows.extend(batch)
        if len(batch) < page_size:
            break
        offset += page_size
    return rows


def extract_brands_from_text(value: object) -> set[str]:
    text = normalize_text(value).lower()
    brands = {brand for needle, brand in BRAND_ALIAS_MAP.items() if needle in text}
    return brands


def parse_pid_from_text(value: object) -> str:
    text = normalize_text(value)
    if not text:
        return ""
    if text.isdigit():
        return text
    import re

    patterns = [r"/product/(\d+)", r"[?&](?:pid|product_id)=([0-9]+)", r"\b(\d{12,})\b"]
    for pattern in patterns:
        match = re.search(pattern, text, flags=re.IGNORECASE)
        if match:
            return match.group(1)
    return ""


def sort_brands(brands: set[str]) -> str:
    order = {"LetMe": 0, "Sparco": 1, "ICYEE": 2, "Stypro": 3}
    return " / ".join(sorted(brands, key=lambda item: (order.get(item, 999), item)))


def score_fact_row(row: dict[str, object]) -> tuple[int, float]:
    score = 0
    for header in [
        "达人名称",
        "达人昵称",
        "主页链接",
        "店铺标签",
        "最近统计日期",
        "最近视频发布时间",
        "合作表_最近合作日期",
        "合作表_当前合作状态",
    ]:
        if normalize_text(row.get(header)):
            score += 1
    for header in ["累计GMV", "近90天GMV", "近30天GMV"]:
        if normalize_number(row.get(header)) > 0:
            score += 2
    return score, normalize_number(row.get("累计GMV"))


def build_creator_mapping(workbook) -> tuple[dict[str, dict[str, object]], dict[str, dict[str, object]]]:
    rows = rows_as_dicts(workbook["Creator映射"], start_row=2)
    by_source: dict[str, dict[str, object]] = {}
    by_existing: dict[str, dict[str, object]] = {}
    for row in rows:
        source_key = normalize_text(row.get("统一达人键"))
        existing_id = normalize_text(row.get("现有达人ID"))
        if source_key and source_key != "说明":
            by_source[source_key] = row
        if existing_id:
            by_existing[existing_id] = row
    return by_source, by_existing


def build_focus_seed_rows(workbook) -> list[dict[str, object]]:
    if CREATOR_POOL_JSON.exists():
        payload = json.loads(CREATOR_POOL_JSON.read_text(encoding="utf-8"))
        rows: list[dict[str, object]] = []
        for creator in payload.get("creators", []):
            rows.append(
                {
                    "达人ID": normalize_text(creator.get("kolId")),
                    "达人名称": normalize_text(creator.get("达人昵称")) or normalize_text(creator.get("kolId")),
                    "平台": normalize_text(creator.get("平台")) or "TikTok",
                    "粉丝量": "",
                    "达人分层(L0/L1/L2/L3)": "",
                    "达人类型": normalize_text(creator.get("内容一级标签")),
                    "备注": normalize_text(creator.get("备注")),
                }
            )
        if rows:
            return rows

    sheet = workbook["达人总览"]
    headers = sheet_headers(sheet)
    rows = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        record = {headers[index]: row[index] for index in range(min(len(headers), len(row)))}
        if normalize_text(record.get("达人ID")):
            rows.append(record)
    return rows


def load_existing_overview_rows(workbook) -> list[dict[str, object]]:
    sheet = workbook["达人总览"]
    headers = sheet_headers(sheet)
    rows: list[dict[str, object]] = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        record = {headers[index]: row[index] for index in range(min(len(headers), len(row)))}
        creator_id = normalize_text(record.get("达人ID"))
        if not creator_id or creator_id == "说明":
            continue
        rows.append(
            {
                header: record.get(header, "")
                for header in OVERVIEW_HEADERS
            }
            | {
                "主页链接": normalize_text(record.get("主页链接")) or f"https://www.tiktok.com/@{creator_id}",
                "统一达人键": normalize_text(record.get("统一达人键")) or creator_id,
            }
        )
    return rows


def load_level_mapping() -> dict[str, str]:
    if not LEVEL_SHEET_PATH.exists():
        return {}
    workbook = load_workbook(LEVEL_SHEET_PATH, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    rows = rows_as_dicts(sheet, start_row=2)
    mapping: dict[str, str] = {}
    for row in rows:
        level = normalize_text(row.get("level"))
        creator_ref = normalize_text(row.get("sparco"))
        key = normalize_handle(creator_ref)
        if key and level:
            mapping[key] = level
    return mapping


def build_raw_creator_meta(workbook) -> tuple[dict[str, dict[str, object]], set[str]]:
    rows = rows_as_dicts(workbook["原始达人表现"], start_row=2)
    meta_by_key: dict[str, dict[str, object]] = {}
    valid_keys: set[str] = set()
    for row in rows:
        key = normalize_text(row.get("统一达人键"))
        if not key or key == "说明":
            continue
        valid_keys.add(key)
        current = meta_by_key.get(key, {})
        candidate_score = sum(
            1
            for header in ["达人名称", "达人主页标识", "统计日期", "店铺", "Affiliate-attributed GMV", "Affiliate followers"]
            if normalize_text(row.get(header))
        )
        current_score = sum(
            1
            for header in ["达人名称", "达人主页标识", "统计日期", "店铺", "Affiliate-attributed GMV", "Affiliate followers"]
            if normalize_text(current.get(header))
        )
        if candidate_score >= current_score:
            meta_by_key[key] = row
    return meta_by_key, valid_keys


def build_db_creator_rollup() -> tuple[dict[str, dict[str, object]], dict[str, dict[str, object]], set[str]]:
    ninety_range = quote("2025-12-26至2026-03-24", safe="")
    baseline_rows = fetch_supabase_rows(
        "tiktok_creator_performance_raw",
        CREATOR_RAW_COLUMNS,
        extra_query=f"stat_date=eq.2026-03-24&range_label=eq.{ninety_range}",
    )
    single_day_label = quote("*单日", safe="*")
    daily_rows = fetch_supabase_rows(
        "tiktok_creator_performance_raw",
        CREATOR_RAW_COLUMNS,
        extra_query=f"range_label=like.{single_day_label}",
    )
    raw_rows = baseline_rows + daily_rows
    if not raw_rows:
        raw_rows = fetch_supabase_rows("tiktok_creator_performance_raw", CREATOR_RAW_COLUMNS)

    deduped_rows_by_store_date_key: dict[tuple[str, str, str], dict[str, object]] = {}
    for row in raw_rows:
        store_tag = normalize_text(row.get("store_tag"))
        stat_date_text = normalize_text(row.get("stat_date"))
        key = normalize_text(row.get("creator_key"))
        if not store_tag or not stat_date_text or not key:
            continue
        dedupe_key = (store_tag, stat_date_text, key)
        existing = deduped_rows_by_store_date_key.get(dedupe_key)
        if existing is None:
            deduped_rows_by_store_date_key[dedupe_key] = row
            continue
        candidate_score = (
            parse_range_span_days(row.get("range_label")),
            normalize_number(row.get("affiliate_gmv")),
            normalize_number(row.get("videos")),
        )
        existing_score = (
            parse_range_span_days(existing.get("range_label")),
            normalize_number(existing.get("affiliate_gmv")),
            normalize_number(existing.get("videos")),
        )
        if candidate_score >= existing_score:
            deduped_rows_by_store_date_key[dedupe_key] = row

    rows = list(deduped_rows_by_store_date_key.values())
    grouped: dict[str, dict[str, object]] = {}
    meta_by_key: dict[str, dict[str, object]] = {}
    followers_by_key: dict[str, tuple[datetime | None, str]] = {}
    best_snapshot_by_key: dict[str, tuple[int, datetime | None, float, float, float]] = {}
    post_snapshot_daily_by_key: dict[str, tuple[float, float, float]] = {}
    valid_keys: set[str] = set()

    for row in rows:
        key = normalize_text(row.get("creator_key"))
        if not key:
            continue
        valid_keys.add(key)
        stat_date = parse_date(row.get("stat_date"))
        gmv = normalize_number(row.get("affiliate_gmv"))
        orders = normalize_number(row.get("attributed_orders"))
        videos = normalize_number(row.get("videos"))
        followers_text = normalize_text(row.get("affiliate_followers"))
        span_days = parse_range_span_days(row.get("range_label"))
        group = grouped.setdefault(
            key,
            {
                "历史总GMV": 0.0,
                "90天GMV": 0.0,
                "历史基线GMV": 0.0,
                "后基线增量GMV": 0.0,
                "累计订单": 0.0,
                "近90天订单": 0.0,
                "累计视频数": 0.0,
                "近90天视频数": 0.0,
                "店铺标签": set(),
                "最近统计日期": "",
            },
        )
        group["店铺标签"].add(normalize_text(row.get("store_tag")))

        if span_days > 1:
            existing = best_snapshot_by_key.get(key)
            snapshot_score = (span_days, stat_date or datetime.min, gmv, orders, videos)
            if existing is None or snapshot_score >= existing:
                best_snapshot_by_key[key] = (span_days, stat_date, gmv, orders, videos)
        else:
            prev = post_snapshot_daily_by_key.get(key, (0.0, 0.0, 0.0))
            post_snapshot_daily_by_key[key] = (prev[0] + gmv, prev[1] + orders, prev[2] + videos)

        if followers_text:
            existing_followers = followers_by_key.get(key)
            if existing_followers is None or ((stat_date or datetime.min) >= (existing_followers[0] or datetime.min)):
                followers_by_key[key] = (stat_date, followers_text)
        if stat_date:
            date_text = stat_date.strftime("%Y-%m-%d")
            if date_text > normalize_text(group.get("最近统计日期")):
                group["最近统计日期"] = date_text
                handle = normalize_text(row.get("creator_handle")) or key
                meta_by_key[key] = {
                    "达人名称": normalize_text(row.get("creator_name")),
                    "达人主页标识": handle,
                    "主页链接": f"https://www.tiktok.com/@{handle.lstrip('@')}" if handle else "",
                    "Affiliate followers": followers_text,
                    "最近统计日期": date_text,
                }

    for key in valid_keys:
        group = grouped.setdefault(key, {"店铺标签": set(), "最近统计日期": ""})
        snapshot = best_snapshot_by_key.get(key)
        daily = post_snapshot_daily_by_key.get(key, (0.0, 0.0, 0.0))
        if snapshot is not None:
            _span, _date, snap_gmv, snap_orders, snap_videos = snapshot
            group["历史基线GMV"] = snap_gmv
            group["后基线增量GMV"] = daily[0]
            group["90天GMV"] = snap_gmv
            group["近90天订单"] = snap_orders
            group["近90天视频数"] = snap_videos
            group["历史总GMV"] = snap_gmv + daily[0]
            group["累计订单"] = snap_orders + daily[1]
            group["累计视频数"] = snap_videos + daily[2]
        else:
            group["历史基线GMV"] = 0.0
            group["后基线增量GMV"] = daily[0]
            group["90天GMV"] = daily[0]
            group["近90天订单"] = daily[1]
            group["近90天视频数"] = daily[2]
            group["历史总GMV"] = daily[0]
            group["累计订单"] = daily[1]
            group["累计视频数"] = daily[2]

    for key, (_followers_date, followers_text) in followers_by_key.items():
        meta = meta_by_key.setdefault(
            key,
            {
                "达人名称": "",
                "达人主页标识": key,
                "主页链接": f"https://www.tiktok.com/@{key.lstrip('@')}" if key else "",
                "最近统计日期": "",
            },
        )
        if followers_text and not normalize_text(meta.get("Affiliate followers")):
            meta["Affiliate followers"] = followers_text
    return grouped, meta_by_key, valid_keys


def build_db_video_rollup() -> tuple[dict[str, dict[str, object]], dict[str, set[str]], dict[str, set[str]], set[str]]:
    rows = fetch_supabase_rows("tiktok_video_performance_raw", VIDEO_RAW_COLUMNS)
    latest_reference = max(
        (
            parse_date(row.get("post_date")) or parse_date(row.get("stat_date"))
            for row in rows
            if parse_date(row.get("post_date")) or parse_date(row.get("stat_date"))
        ),
        default=datetime.now(),
    )
    cutoff_30 = latest_reference - timedelta(days=29)
    cutoff_90 = latest_reference - timedelta(days=89)
    grouped: dict[str, dict[str, object]] = {}
    brands_by_key: dict[str, set[str]] = {}
    brands_by_pid: dict[str, set[str]] = {}
    valid_keys: set[str] = set()
    for row in rows:
        key = normalize_text(row.get("creator_key"))
        if not key:
            continue
        valid_keys.add(key)
        post_date = parse_date(row.get("post_date"))
        stat_date = parse_date(row.get("stat_date"))
        gmv = normalize_number(row.get("affiliate_video_gmv"))
        orders = normalize_number(row.get("video_orders"))
        brand = canonical_brand(row.get("store_tag"))
        if brand:
            brands_by_key.setdefault(key, set()).add(brand)
            pid = normalize_int_string(row.get("product_id"))
            if pid:
                brands_by_pid.setdefault(pid, set()).add(brand)
        group = grouped.setdefault(
            key,
            {
                "近30天发布视频GMV": 0.0,
                "近30天订单": 0.0,
                "近30天视频数": 0.0,
                "近90天GMV": 0.0,
                "近90天订单": 0.0,
                "近90天视频数": 0.0,
                "累计GMV": 0.0,
                "累计订单": 0.0,
                "累计视频数": 0.0,
                "最近视频发布时间": "",
                "最近统计日期": "",
            },
        )
        group["累计GMV"] += gmv
        group["累计订单"] += orders
        group["累计视频数"] += 1
        if stat_date and stat_date.strftime("%Y-%m-%d") > normalize_text(group.get("最近统计日期")):
            group["最近统计日期"] = stat_date.strftime("%Y-%m-%d")
        if post_date and post_date.strftime("%Y-%m-%d") > normalize_text(group.get("最近视频发布时间")):
            group["最近视频发布时间"] = post_date.strftime("%Y-%m-%d")
        if post_date and post_date >= cutoff_30:
            group["近30天发布视频GMV"] += gmv
            group["近30天订单"] += orders
            group["近30天视频数"] += 1
        if post_date and post_date >= cutoff_90:
            group["近90天GMV"] += gmv
            group["近90天订单"] += orders
            group["近90天视频数"] += 1
    return grouped, brands_by_key, brands_by_pid, valid_keys


def build_video_brand_maps(workbook) -> tuple[dict[str, set[str]], dict[str, set[str]]]:
    rows = rows_as_dicts(workbook["原始视频表现"], start_row=2)
    brands_by_key: dict[str, set[str]] = {}
    brands_by_pid: dict[str, set[str]] = {}
    for row in rows:
        key = normalize_text(row.get("统一达人键"))
        if not key or key == "说明":
            continue
        brand = canonical_brand(row.get("店铺"))
        if not brand:
            continue
        brands_by_key.setdefault(key, set()).add(brand)
        pid = normalize_int_string(row.get("Product ID"))
        if pid:
            brands_by_pid.setdefault(pid, set()).add(brand)
    return brands_by_key, brands_by_pid


def build_fact_rows(workbook) -> dict[str, dict[str, object]]:
    rows = rows_as_dicts(workbook["统一达人事实表"], start_row=2)
    fact_by_key: dict[str, dict[str, object]] = {}
    for row in rows:
        key = normalize_text(row.get("统一达人键"))
        if not key or key == "说明":
            continue
        current = fact_by_key.get(key)
        if current is None or score_fact_row(row) > score_fact_row(current):
            fact_by_key[key] = row
    return fact_by_key


def build_records(workbook) -> list[dict[str, object]]:
    rows = rows_as_dicts(workbook["合作记录明细"], start_row=2)
    result = []
    for row in rows:
        creator_id = normalize_text(row.get("达人ID"))
        if not creator_id:
            continue
        result.append({header: row.get(header, "") for header in RECORD_HEADERS})
    return result


def compute_timeout(near30_video_gmv: float, latest_fee: float, recent_coop_date: str, publish_gap: object) -> str:
    recent_dt = parse_date(recent_coop_date)
    if not recent_dt or latest_fee <= 0 or near30_video_gmv <= latest_fee:
        return "N"
    deadline = recent_dt + timedelta(days=7)
    now_dt = datetime.now()
    gap_text = normalize_text(publish_gap)
    if gap_text:
        try:
            latest_post_date = now_dt - timedelta(days=int(float(gap_text)))
            return "Y" if latest_post_date > deadline and now_dt > deadline else "N"
        except ValueError:
            return "Y" if now_dt > deadline else "N"
    return "Y" if now_dt > deadline else "N"


def load_pipeline_state() -> dict[str, object]:
    if not PIPELINE_STATE_PATH.exists():
        return {}
    try:
        return json.loads(PIPELINE_STATE_PATH.read_text(encoding="utf-8"))
    except json.JSONDecodeError:
        return {}


def load_automation_status() -> dict[str, str]:
    if not AUTOMATION_TOML_PATH.exists():
        return {"status": "MISSING", "schedule": ""}
    raw_text = AUTOMATION_TOML_PATH.read_text(encoding="utf-8")
    if tomllib:
        try:
            data = tomllib.loads(raw_text)
        except tomllib.TOMLDecodeError:
            return {"status": "BROKEN", "schedule": ""}
    else:
        data: dict[str, str] = {}
        for line in raw_text.splitlines():
            if "=" not in line:
                continue
            key, value = line.split("=", 1)
            data[key.strip()] = value.strip().strip('"')
    return {
        "status": normalize_text(data.get("status")) or "UNKNOWN",
        "schedule": normalize_text(data.get("rrule")),
    }


def store_health_row(store_key: str, store: dict[str, object]) -> dict[str, object]:
    status = normalize_text(store.get("last_status") or store.get("last_mode")) or "未记录"
    window_status = normalize_text(store.get("window_status"))
    last_daily = normalize_text(store.get("last_daily_date"))
    next_date = normalize_text(store.get("next_increment_date"))
    last_error = normalize_text(store.get("last_error"))
    if status == "success" and last_daily:
        value = f"已同步至 {last_daily}"
    elif status == "skipped":
        value = "已跳过"
    elif status == "error":
        value = "运行失败"
    elif status == "bootstrap_required":
        value = "待导入历史基线"
    elif status == "waiting" and last_daily:
        value = f"已同步至 {last_daily}"
        window_status = "waiting"
    else:
        value = status
    if window_status == "waiting" and next_date:
        note = f"等待 {next_date} 数据窗口"
    elif next_date:
        note = f"待同步日期 {next_date}"
    else:
        note = "无待同步日期"
    return {
        "store": store_key.upper(),
        "value": value,
        "note": note,
        "error": last_error,
    }


def status_display(status: object, kind: str) -> tuple[str, str]:
    text = normalize_text(status)
    if kind == "data_sync":
        mapping = {
            "success": ("采集入库成功", "4店数据已采集并完成入库"),
            "partial_success": ("采集部分成功", "存在跳过或待基线店铺"),
            "partial_failure": ("采集部分失败", "部分店铺或入库步骤失败"),
            "failed": ("采集入库失败", "本轮采集或数据库同步未完成"),
            "running": ("采集入库进行中", "正在执行导出、标准化或入库"),
            "waiting": ("等待数据窗口", "当前没有 backlog，等待下一天数据可拉取"),
        }
    else:
        mapping = {
            "success": ("页面重建成功", "工作簿和页面快照已刷新"),
            "failed": ("页面重建失败", "后处理未完成，请查看失败原因"),
            "running": ("页面重建中", "正在生成工作簿或页面快照"),
            "pending": ("页面重建待开始", "等待采集入库完成后继续"),
            "not_started": ("页面重建未开始", "采集未完成，暂未进入重建"),
            "not_needed": ("页面重建未执行", "本轮无新数据，无需刷新页面快照"),
        }
    return mapping.get(text, ("未记录", "暂无状态记录"))


def build_sync_health() -> dict[str, object]:
    pipeline_state = load_pipeline_state()
    stores = pipeline_state.get("stores", {}) if isinstance(pipeline_state, dict) else {}
    run = pipeline_state.get("last_pipeline_run", {}) if isinstance(pipeline_state, dict) else {}
    automation = load_automation_status()
    failures = run.get("failures", []) if isinstance(run, dict) else []
    last_failure = ""
    if isinstance(failures, list) and failures:
        latest = failures[0]
        if isinstance(latest, dict):
            last_failure = f"{normalize_text(latest.get('store'))}: {normalize_text(latest.get('message'))}"
    if not last_failure:
        for store_key in ("letme", "stypro", "sparco", "icyee"):
            store = stores.get(store_key, {}) if isinstance(stores, dict) else {}
            last_error = normalize_text(store.get("last_error"))
            if last_error and normalize_text(store.get("last_status")) == "error":
                last_failure = f"{store_key.upper()}: {last_error}"
                break

    store_rows = []
    success_count = 0
    waiting_count = 0
    failed_count = 0
    latest_data_date = ""
    synced_dates: list[str] = []
    for store_key in ("letme", "stypro", "sparco", "icyee"):
        store = stores.get(store_key, {}) if isinstance(stores, dict) else {}
        row = store_health_row(store_key, store)
        store_rows.append(row)
        value = normalize_text(row.get("value"))
        note = normalize_text(row.get("note"))
        last_daily = normalize_text(store.get("last_daily_date"))
        if last_daily:
            synced_dates.append(last_daily)
            if not latest_data_date or last_daily > latest_data_date:
                latest_data_date = last_daily
        if "已同步至" in value:
            success_count += 1
        if "等待" in note:
            waiting_count += 1
        if "失败" in value:
            failed_count += 1
    all_stores_synced_through = min(synced_dates) if synced_dates else ""

    last_success_run_at = normalize_text(run.get("last_success_run_at"))
    if not last_success_run_at:
        store_successes = [
            normalize_text(store.get("last_success_at"))
            for store in stores.values()
            if isinstance(store, dict) and normalize_text(store.get("last_success_at"))
        ]
        last_success_run_at = max(store_successes) if store_successes else "未记录"

    data_sync_status = normalize_text(run.get("data_sync_status")) or normalize_text(run.get("overall_status"))
    rebuild_status = normalize_text(run.get("rebuild_status"))
    data_sync_value, data_sync_note = status_display(data_sync_status, "data_sync")
    rebuild_value, rebuild_note = status_display(rebuild_status, "rebuild")
    db_sync_error = normalize_text(run.get("db_sync_error"))
    rebuild_error = normalize_text(run.get("rebuild_error"))
    if db_sync_error:
        data_sync_note = f"{data_sync_note} · {db_sync_error}"
    if rebuild_error:
        rebuild_note = f"{rebuild_note} · {rebuild_error}"

    return {
        "automationStatus": automation.get("status", "UNKNOWN"),
        "schedule": automation.get("schedule", ""),
        "lastRunAt": normalize_text(run.get("run_at")) or "未记录",
        "lastSuccessRunAt": last_success_run_at,
        "latestDataDate": latest_data_date or "未记录",
        "allStoresSyncedThrough": all_stores_synced_through or "未记录",
        "nextSyncDate": normalize_text(run.get("next_sync_date")) or "未记录",
        "dataSyncStatus": data_sync_value,
        "dataSyncNote": data_sync_note,
        "rebuildStatus": rebuild_value,
        "rebuildNote": rebuild_note,
        "summary": f"{success_count} 已同步 / {waiting_count} 等待 / {failed_count} 失败",
        "lastFailure": last_failure or "无",
        "stores": store_rows,
    }


def build_sync_health_metrics(sync_health: dict[str, object]) -> list[dict[str, object]]:
    metrics: list[dict[str, object]] = [
        {
            "指标": "自动化任务状态",
            "数值": sync_health.get("automationStatus", "UNKNOWN"),
            "说明": "当前正式日更任务定义",
            "": sync_health.get("schedule", ""),
        },
        {
            "指标": "上次自动化运行",
            "数值": sync_health.get("lastRunAt", "未记录"),
            "说明": "最近一次流水线运行日期",
            "": "",
        },
        {
            "指标": "上次成功同步",
            "数值": sync_health.get("lastSuccessRunAt", "未记录"),
            "说明": "最近一次成功完成日更的日期",
            "": "",
        },
        {
            "指标": "最新单店更新到",
            "数值": sync_health.get("latestDataDate", "未记录"),
            "说明": "四店中任一店铺已更新到的最新日期",
            "": "",
        },
        {
            "指标": "四店统一更新到",
            "数值": sync_health.get("allStoresSyncedThrough", "未记录"),
            "说明": "当前四店都已覆盖到的共同最新日期",
            "": "",
        },
        {
            "指标": "采集入库状态",
            "数值": sync_health.get("dataSyncStatus", "未记录"),
            "说明": "浏览器拉起、导出、标准化和数据库入库状态",
            "": sync_health.get("dataSyncNote", ""),
        },
        {
            "指标": "页面重建状态",
            "数值": sync_health.get("rebuildStatus", "未记录"),
            "说明": "工作簿与页面快照重建状态",
            "": sync_health.get("rebuildNote", ""),
        },
        {
            "指标": "当前待同步日期",
            "数值": sync_health.get("nextSyncDate", "未记录"),
            "说明": "四店铺中最早待推进的增量日期",
            "": "",
        },
        {
            "指标": "店铺同步结果",
            "数值": sync_health.get("summary", ""),
            "说明": "最近一次自动化四店铺结果",
            "": "",
        },
        {
            "指标": "最近失败原因",
            "数值": sync_health.get("lastFailure", "无"),
            "说明": "如果最近一次没有失败，这里显示无",
            "": "",
        },
    ]
    for store in sync_health.get("stores", []):
        metrics.append(
            {
                "指标": f"{normalize_text(store.get('store'))}同步状态",
                "数值": normalize_text(store.get("value")) or "未记录",
                "说明": normalize_text(store.get("note")) or "无待同步日期",
                "": normalize_text(store.get("error")),
            }
        )
    return metrics


def build_brand_tags(
    key: str,
    fact_row: dict[str, object],
    video_brands_by_key: dict[str, set[str]],
    brands_by_pid: dict[str, set[str]],
) -> str:
    brands = set(video_brands_by_key.get(key, set()))
    manual_link = normalize_text(fact_row.get("手工_复投产品链接"))
    manual_pid = parse_pid_from_text(fact_row.get("手工_复投产品PID")) or parse_pid_from_text(manual_link)
    brands |= extract_brands_from_text(manual_link)
    if manual_pid:
        brands |= brands_by_pid.get(manual_pid, set())
    return sort_brands(brands)


def build_row(
    key: str,
    fact_row: dict[str, object],
    meta_row: dict[str, object],
    mapping_row: dict[str, object],
    focus_seed: dict[str, object] | None,
    brand_tags: str,
    level_mapping: dict[str, str],
) -> dict[str, object]:
    creator_id = (
        normalize_text((mapping_row or {}).get("现有达人ID"))
        or normalize_text((meta_row or {}).get("达人主页标识"))
        or normalize_text((focus_seed or {}).get("达人ID"))
        or key
    )
    creator_name = (
        normalize_text((focus_seed or {}).get("达人名称"))
        or normalize_text((mapping_row or {}).get("达人名称_现有库"))
        or normalize_text(fact_row.get("达人名称"))
        or normalize_text((meta_row or {}).get("达人名称"))
        or creator_id
    )
    level = (
        normalize_text((focus_seed or {}).get("达人分层(L0/L1/L2/L3)"))
        or level_mapping.get(normalize_handle(creator_id))
        or level_mapping.get(normalize_handle(creator_name))
        or level_mapping.get(normalize_handle(key))
    )
    history_gmv = normalize_number(fact_row.get("累计GMV")) or normalize_number((focus_seed or {}).get("历史总GMV"))
    gm90 = normalize_number(fact_row.get("近90天GMV")) or normalize_number((focus_seed or {}).get("90天GMV"))
    gm30 = normalize_number(fact_row.get("近30天GMV")) or normalize_number((focus_seed or {}).get("近30天发布视频gmv"))
    recent_coop_date = normalize_text(fact_row.get("合作表_最近合作日期")) or normalize_text((focus_seed or {}).get("最近合作日期"))
    current_status = normalize_text(fact_row.get("合作表_当前合作状态")) or normalize_text((focus_seed or {}).get("当前合作状态"))
    coop30 = normalize_text(fact_row.get("合作表_近30天合作次数")) or normalize_text((focus_seed or {}).get("近30天合作次数"))
    interval_days = normalize_text(fact_row.get("合作表_平均间隔天数")) or normalize_text((focus_seed or {}).get("平均间隔天数"))
    publish_gap = normalize_text(fact_row.get("合作表_距离上次发布天数")) or normalize_text((focus_seed or {}).get("距离上次发布天数"))
    delivery_days = normalize_text(fact_row.get("合作表_平均交付时间")) or normalize_text((focus_seed or {}).get("平均交付时间"))
    latest_fee = normalize_number(fact_row.get("合作表_最近合作费用"))
    repurchase = normalize_text((focus_seed or {}).get("是否进入复投(Y/N)"))
    if not repurchase:
        repurchase = "Y" if gm30 > latest_fee else "N"
    timeout = normalize_text((focus_seed or {}).get("是否超时未合作"))
    if not timeout:
        timeout = compute_timeout(gm30, latest_fee, recent_coop_date, publish_gap)
    note = normalize_text((focus_seed or {}).get("备注")) or normalize_text(fact_row.get("备注"))
    row = {
        "达人ID": creator_id,
        "达人名称": creator_name,
        "平台": normalize_text((focus_seed or {}).get("平台")) or "TikTok",
        "粉丝量": normalize_text((focus_seed or {}).get("粉丝量")) or normalize_text((meta_row or {}).get("Affiliate followers")),
        "达人分层(L0/L1/L2/L3)": level,
        "达人类型": normalize_text((focus_seed or {}).get("达人类型")),
        "品牌标签": brand_tags,
        "历史总GMV": round(history_gmv, 2),
        "90天GMV": round(gm90, 2),
        "近30天发布视频gmv": round(gm30, 2),
        "最近合作日期": recent_coop_date,
        "当前合作状态": current_status,
        "近30天合作次数": coop30,
        "平均间隔天数": interval_days,
        "距离上次发布天数": publish_gap,
        "近30天是否出单(Y/N)": "Y" if gm30 > 0 else "N",
        "是否进入复投(Y/N)": repurchase,
        "是否超时未合作": timeout,
        "平均交付时间": delivery_days,
        "优先级": normalize_text((focus_seed or {}).get("优先级")),
        "下一步动作": normalize_text((focus_seed or {}).get("下一步动作")),
        "负责人": normalize_text((focus_seed or {}).get("负责人")),
        "截止日期": normalize_text((focus_seed or {}).get("截止日期")),
        "备注": note,
        "主页链接": normalize_text(fact_row.get("主页链接")) or f"https://www.tiktok.com/@{creator_id}",
        "统一达人键": key,
    }
    return row


FACT_ENRICHMENT_FIELDS = {
    "合作表_最近合作日期",
    "合作表_当前合作状态",
    "合作表_近30天合作次数",
    "合作表_平均间隔天数",
    "合作表_距离上次发布天数",
    "合作表_平均交付时间",
    "合作表_最近合作费用",
    "手工_复投产品链接",
    "手工_复投产品PID",
    "备注",
}


def build_payload() -> tuple[dict[str, object], list[dict[str, object]]]:
    workbook = load_workbook(WORKBOOK_PATH, read_only=True, data_only=True)
    level_mapping = load_level_mapping()
    focus_seed_rows = build_focus_seed_rows(workbook)
    creator_mapping_by_source, creator_mapping_by_existing = build_creator_mapping(workbook)
    workbook_fact_by_key = build_fact_rows(workbook)
    record_rows = build_records(workbook)
    source_label = str(WORKBOOK_PATH)

    try:
        creator_rollup, creator_meta_by_key, creator_keys = build_db_creator_rollup()
        video_rollup, video_brands_by_key, brands_by_pid, video_keys = build_db_video_rollup()
        valid_keys = creator_keys | video_keys
        fact_by_key: dict[str, dict[str, object]] = {}
        for key in sorted(valid_keys):
            workbook_fact = workbook_fact_by_key.get(key, {})
            merged = {field: workbook_fact.get(field, "") for field in FACT_ENRICHMENT_FIELDS}
            creator_stats = creator_rollup.get(key, {})
            video_stats = video_rollup.get(key, {})
            if creator_stats:
                history_base = normalize_number(creator_stats.get("历史基线GMV"))
                history_delta = normalize_number(creator_stats.get("后基线增量GMV"))
                merged["累计GMV"] = round(history_base + history_delta, 2)
                merged["近90天GMV"] = round(history_base or normalize_number(creator_stats.get("90天GMV")), 2)
                merged["最近统计日期"] = normalize_text(creator_stats.get("最近统计日期"))
                merged["达人名称"] = normalize_text(creator_meta_by_key.get(key, {}).get("达人名称")) or normalize_text(merged.get("达人名称"))
                merged["主页链接"] = normalize_text(creator_meta_by_key.get(key, {}).get("主页链接")) or normalize_text(merged.get("主页链接"))
            else:
                merged["累计GMV"] = 0
                merged["近90天GMV"] = 0
            if video_stats:
                merged["近30天GMV"] = round(normalize_number(video_stats.get("近30天发布视频GMV")), 2)
                merged["最近视频发布时间"] = normalize_text(video_stats.get("最近视频发布时间"))
                merged["最近统计日期"] = normalize_text(video_stats.get("最近统计日期")) or normalize_text(merged.get("最近统计日期"))
            else:
                merged["近30天GMV"] = 0
                merged["最近视频发布时间"] = ""
            fact_by_key[key] = merged
        source_label = "supabase_raw + workbook_enrichment"
    except Exception:
        creator_meta_by_key, valid_keys = build_raw_creator_meta(workbook)
        fact_by_key = workbook_fact_by_key
        video_brands_by_key, brands_by_pid = build_video_brand_maps(workbook)
        source_label = f"{WORKBOOK_PATH} (fallback)"

    overview_rows: list[dict[str, object]] = []
    all_keys = sorted(set(fact_by_key) | set(valid_keys))
    for key in all_keys:
        fact_row = fact_by_key.get(key, {})
        meta_row = creator_meta_by_key.get(key, {})
        mapping_row = creator_mapping_by_source.get(key, {})
        brand_tags = build_brand_tags(key, fact_row, video_brands_by_key, brands_by_pid)
        overview_rows.append(build_row(key, fact_row, meta_row, mapping_row, None, brand_tags, level_mapping))

    focus_rows: list[dict[str, object]] = []
    missing_focus_ids: list[str] = []
    overview_by_id = {normalize_text(row.get("达人ID")): row for row in overview_rows}
    for seed in focus_seed_rows:
        creator_id = normalize_text(seed.get("达人ID"))
        mapping_row = creator_mapping_by_existing.get(creator_id, {})
        key = normalize_text(mapping_row.get("统一达人键")) or creator_id
        fact_row = fact_by_key.get(key, {})
        meta_row = creator_meta_by_key.get(key, {})
        brand_tags = build_brand_tags(key, fact_row, video_brands_by_key, brands_by_pid)
        focus_row = build_row(key, fact_row, meta_row, mapping_row, seed, brand_tags, level_mapping)
        if creator_id not in overview_by_id:
            missing_focus_ids.append(creator_id)
        focus_rows.append(focus_row)

    gmv_focus_rows = [row for row in overview_rows if normalize_number(row.get("90天GMV")) > 0]

    assumptions = [
        "全量池继续保留在后台作为 Creator 全量唯一达人聚合真源，但前台不再直接展示全量明细。",
        "GMV重点池按 90天GMV > 0 自动生成，每次日更重建后会自动重新判定是否入池。",
        "品牌标签为独立运营字段，当前按历史视频实际带货的店铺品牌累计生成，并补充手工产品链接/PID可识别出的品牌。",
        "核心达人类型继续沿用既有 189 人标签结果；未命中全量 Creator 的达人类型允许为空，不自动补打新标签。",
        "GMV 与合作公式继续沿用当前同步版工作簿口径，本轮不重算业务规则。",
    ]

    stats = {
        "overviewCount": len(overview_rows),
        "gmvFocusCount": len(gmv_focus_rows),
        "focusCount": len(focus_rows),
        "recordCount": len(record_rows),
        "highPriorityCount": sum(1 for row in focus_rows if normalize_text(row.get("优先级")) == "高"),
        "missingFocusCount": len(missing_focus_ids),
    }

    sync_health = build_sync_health()
    metrics = rows_as_dicts(workbook["运营驾驶舱"], start_row=2) + build_sync_health_metrics(sync_health)

    payload = {
        "generatedAt": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "source": source_label,
        "assumptions": assumptions,
        "stats": stats,
        "syncHealth": sync_health,
        "overview": [],
        "gmvFocusPool": [{header: row.get(header, "") for header in OVERVIEW_HEADERS} | {"主页链接": row.get("主页链接", ""), "统一达人键": row.get("统一达人键", "")} for row in gmv_focus_rows],
        "focusPool": [{header: row.get(header, "") for header in FOCUS_HEADERS} | {"主页链接": row.get("主页链接", ""), "统一达人键": row.get("统一达人键", "")} for row in focus_rows],
        "records": record_rows,
        "metrics": metrics,
        "missingFocusIds": missing_focus_ids,
    }
    return payload, overview_rows


def export_csv(overview_rows: list[dict[str, object]]) -> None:
    with OUTPUT_CSV.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=OVERVIEW_HEADERS)
        writer.writeheader()
        for row in overview_rows:
            writer.writerow({header: row.get(header, "") for header in OVERVIEW_HEADERS})


def main() -> None:
    payload, overview_rows = build_payload()
    OUTPUT_JSON.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    OUTPUT_JS.write_text(
        "window.__CORE_CREATOR_DASHBOARD__ = "
        + json.dumps(payload, ensure_ascii=False, indent=2)
        + ";\n",
        encoding="utf-8",
    )
    export_csv(overview_rows)
    print(f"Wrote {OUTPUT_JSON}")
    print(f"Wrote {OUTPUT_JS}")
    print(f"Wrote {OUTPUT_CSV}")


if __name__ == "__main__":
    main()
