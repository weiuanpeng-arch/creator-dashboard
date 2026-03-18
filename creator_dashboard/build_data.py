from __future__ import annotations

import json
from datetime import datetime
from pathlib import Path
from urllib.error import URLError
from urllib.parse import quote
from urllib.request import Request, urlopen

from openpyxl import load_workbook


BASE_DIR = Path(__file__).resolve().parent
SOURCE_XLSX = Path("/Users/apple/Desktop/达人标签管理_合作3次以上.xlsx")
OUTPUT_JSON = BASE_DIR / "data" / "creator_pool.json"
OUTPUT_JS = BASE_DIR / "data" / "creator_pool.js"
OUTPUT_CSV = BASE_DIR / "data" / "creator_pool.csv"
PUBLIC_READ_SYNC_SETTINGS = {
    "supabase_url": "https://sbznfjnsirajqkkcwayj.supabase.co",
    "anon_key": "sb_publishable_tM67K7Mi1qDUkemhgzDuGg_dsdwitBT",
    "workspace_id": "creator-dashboard-prod",
}


def normalize(value):
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    return str(value).strip()


def load_sheet_rows(workbook_path: Path, sheet_name: str) -> list[dict[str, str]]:
    workbook = load_workbook(workbook_path, data_only=True)
    sheet = workbook[sheet_name]
    header = [normalize(cell.value) for cell in sheet[1]]
    rows = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        record = {}
        for index, key in enumerate(header):
            if key:
                record[key] = normalize(row[index] if index < len(row) else "")
        rows.append(record)
    return rows


def supabase_request(path: str) -> list[dict]:
    request = Request(
        f'{PUBLIC_READ_SYNC_SETTINGS["supabase_url"]}/rest/v1{path}',
        headers={
            "apikey": PUBLIC_READ_SYNC_SETTINGS["anon_key"],
            "Authorization": f'Bearer {PUBLIC_READ_SYNC_SETTINGS["anon_key"]}',
        },
    )
    with urlopen(request, timeout=20) as response:
        return json.loads(response.read().decode("utf-8"))


def load_remote_state() -> tuple[dict[str, dict[str, str]], list[dict[str, str]]]:
    workspace = quote(PUBLIC_READ_SYNC_SETTINGS["workspace_id"], safe="")
    try:
        override_rows = supabase_request(
            f"/creator_sync_overrides?workspace_id=eq.{workspace}&select=kol_id,fields"
        )
        custom_tag_rows = supabase_request(
            f"/creator_sync_tags?workspace_id=eq.{workspace}&select=id,tag_category,tag_dimension,tag_name,brand_scope,definition,created_by,updated_at&order=updated_at.desc"
        )
    except URLError as error:
        print(f"Warning: failed to load remote state from Supabase: {error}")
        return {}, []

    overrides = {}
    for row in override_rows or []:
        fields = row.get("fields") or {}
        overrides[row.get("kol_id", "")] = {key: normalize(value) for key, value in fields.items()}

    custom_tags = []
    for row in custom_tag_rows or []:
        custom_tags.append(
            {
                "标签大类": normalize(row.get("tag_category")),
                "标签维度": normalize(row.get("tag_dimension")),
                "标签名称": normalize(row.get("tag_name")),
                "适配品牌": normalize(row.get("brand_scope")),
                "定义/什么时候打这个标签": normalize(row.get("definition")),
            }
        )

    return overrides, custom_tags


def build_payload() -> dict:
    creators = load_sheet_rows(SOURCE_XLSX, "达人主表_合作3次以上")
    tags = load_sheet_rows(SOURCE_XLSX, "标签表")
    brands = load_sheet_rows(SOURCE_XLSX, "品牌说明")
    remote_overrides, remote_custom_tags = load_remote_state()

    for creator in creators:
        creator.update(remote_overrides.get(creator.get("kolId", ""), {}))
        creator["合作次数"] = int(float(creator["合作次数"] or 0))
        creator["contentTags"] = [
            creator.get("内容一级标签", ""),
            creator.get("内容二级标签", ""),
            creator.get("内容形式标签", ""),
            creator.get("人设/风格标签", ""),
            creator.get("受众标签", ""),
        ]
        creator["contentTags"] = [tag for tag in creator["contentTags"] if tag]
        creator["productTags"] = [
            creator.get("带货一级类目", ""),
            creator.get("带货二级类目", ""),
        ]
        creator["productTags"] = [tag for tag in creator["productTags"] if tag]

    tags.extend(remote_custom_tags)

    payload = {
        "generatedAt": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "source": str(SOURCE_XLSX),
        "stats": {
            "creatorCount": len(creators),
            "brandCount": len(brands),
            "tagCount": len(tags),
            "processedCount": sum(1 for creator in creators if creator.get("是否已打标") in {"已初筛", "已完成"}),
        },
        "brands": brands,
        "tags": tags,
        "creators": creators,
    }
    return payload


def main() -> None:
    payload = build_payload()
    json_text = json.dumps(payload, ensure_ascii=False, indent=2)
    OUTPUT_JSON.write_text(json_text, encoding="utf-8")
    OUTPUT_JS.write_text(f"window.__CREATOR_POOL__ = {json_text};\n", encoding="utf-8")
    csv_headers = [
        "kolId",
        "合作次数",
        "平台",
        "主页链接",
        "达人昵称",
        "首次合作时间",
        "最近合作时间",
        "最近合作状态",
        "最近跟进人",
        "是否复投过",
        "历史合作类型",
        "历史合作SPU",
        "内容一级标签",
        "内容二级标签",
        "内容形式标签",
        "人设/风格标签",
        "受众标签",
        "带货一级类目",
        "带货二级类目",
        "适配品牌",
        "转化形式",
        "合作分层",
        "是否已打标",
        "打标依据链接",
        "备注",
    ]
    csv_lines = [",".join(f'"{header}"' for header in csv_headers)]
    for creator in payload["creators"]:
        line = []
        for header in csv_headers:
            value = str(creator.get(header, "")).replace('"', '""')
            line.append(f'"{value}"')
        csv_lines.append(",".join(line))
    OUTPUT_CSV.write_text("\ufeff" + "\n".join(csv_lines) + "\n", encoding="utf-8")
    print(f"Wrote {OUTPUT_JSON}")
    print(f"Wrote {OUTPUT_JS}")
    print(f"Wrote {OUTPUT_CSV}")


if __name__ == "__main__":
    main()
