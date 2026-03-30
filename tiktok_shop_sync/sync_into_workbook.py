from __future__ import annotations

import csv
import json
import os
import re
import shutil
from collections import defaultdict
from datetime import datetime, timedelta
from difflib import SequenceMatcher
from pathlib import Path
from urllib.error import HTTPError, URLError
from urllib.parse import quote
from urllib.request import Request, urlopen

from openpyxl import Workbook, load_workbook
from build_sync_workbook import SOURCE_WORKBOOK, build_notes_sheet, build_sync_sheets, seed_creator_mapping
from workbook_io import (
    DESKTOP_WORKBOOK_PATH,
    INTERNAL_WORKBOOK_PATH,
    ensure_internal_workbook,
    save_workbook_safely,
)


DEFAULT_WORKBOOK_PATH = DESKTOP_WORKBOOK_PATH
FALLBACK_WORKBOOK_PATH = INTERNAL_WORKBOOK_PATH
VIDEO_CSV = Path("/Users/apple/Documents/Playground/tiktok_shop_sync/data/normalized/video_performance.csv")
CREATOR_CSV = Path("/Users/apple/Documents/Playground/tiktok_shop_sync/data/normalized/creator_performance.csv")
CREATOR_HISTORY_CSV = Path("/Users/apple/Documents/Playground/tiktok_shop_sync/data/normalized/creator_history_gmv.csv")
STATE_PATH = Path("/Users/apple/Documents/Playground/tiktok_shop_sync/data/pipeline_state.json")
COOP_XLSX_NAME = "cooperation_export_1774577590040.xlsx"
COOP_XLSX_CANDIDATES = [
    Path(
        "/Users/apple/Library/Containers/com.tencent.xinWeChat/Data/Documents/xwechat_files/"
        "wxid_2514345143114_b48d/temp/drag/cooperation_export_1774577590040.xlsx"
    ),
    Path(
        "/Users/apple/Library/Containers/com.tencent.xinWeChat/Data/Documents/xwechat_files/"
        "wxid_2514345143114_b48d/msg/file/2026-03/cooperation_export_1774577590040.xlsx"
    ),
]
SKU_XLSX = Path("/Users/apple/Desktop/SKU_COST.xlsx")
CREATOR_POOL_JSON = Path("/Users/apple/Documents/Playground/creator_dashboard/data/creator_pool.json")
LEVEL_SHEET_PATH = Path("/Users/apple/Downloads/达人等级底表/达人等级底表.xlsx")

PUBLIC_READ_SYNC_SETTINGS = {
    "supabase_url": "https://sbznfjnsirajqkkcwayj.supabase.co",
    "anon_key": "sb_publishable_tM67K7Mi1qDUkemhgzDuGg_dsdwitBT",
    "workspace_id": "creator-dashboard-prod",
}
SUPABASE_READ_TIMEOUT_SECONDS = float(os.environ.get("TIKTOK_PUBLIC_READ_TIMEOUT_SECONDS", "3"))

ACTIVE_COOP_STATUSES = {"Pending MOU", "Pending review", "MOU submitted", "Pending publish review"}
CORE_OVERRIDE_MAP = {
    "core_复投产品链接": "复投产品链接",
    "core_复投产品PID": "复投产品PID",
    "core_优先级": "优先级",
    "core_下一步动作": "下一步动作",
    "core_负责人": "负责人",
    "core_截止日期": "截止日期",
    "core_备注": "备注",
}

RAW_VIDEO_HEADERS = [
    "抓取日期",
    "统计日期",
    "店铺",
    "平台",
    "来源",
    "时间范围",
    "达人名称",
    "达人主页标识",
    "统一达人键",
    "Video title",
    "Video ID",
    "Post date",
    "Video link",
    "Product name",
    "Product ID",
    "Affiliate video-attributed GMV",
    "Video-attributed orders",
    "AOV",
    "Avg. GMV per customer",
    "Video-attributed items sold",
    "Refunds",
    "Items refunded",
    "Est. commission",
    "数据批次ID",
    "原文件名",
    "备注",
]

RAW_CREATOR_HEADERS = [
    "抓取日期",
    "统计日期",
    "店铺",
    "平台",
    "来源",
    "时间范围",
    "达人名称",
    "达人主页标识",
    "统一达人键",
    "Affiliate-attributed GMV",
    "Attributed orders",
    "Affiliate-attributed items sold",
    "Refunds",
    "Items refunded",
    "AOV",
    "Est. commission",
    "Videos",
    "LIVE streams",
    "Avg. daily products sold",
    "Affiliate followers",
    "Samples shipped",
    "数据批次ID",
    "原文件名",
    "备注",
]

COOP_MAPPING_HEADERS = [
    "原始kolId",
    "规范化kolId",
    "统一达人键",
    "现有达人ID",
    "达人名称_来源",
    "达人名称_现有库",
    "匹配方式",
    "匹配置信度",
    "是否自动生效",
    "备注",
]

AUTO_CANDIDATE_HEADERS = [
    "原始kolId",
    "候选统一达人键",
    "候选现有达人ID",
    "匹配方式",
    "匹配置信度",
    "是否已采用",
]

PRODUCT_RESULT_HEADERS = [
    "合作ID",
    "统一达人键",
    "现有达人ID",
    "开始时间",
    "结束时间",
    "合作状态",
    "合作费用",
    "合作SPU列表",
    "合作SKU列表",
    "复投产品链接",
    "复投产品PID",
    "匹配视频ID",
    "匹配视频链接",
    "匹配发布时间",
    "视频产品名称",
    "视频产品PID",
    "视频实际产品SKU",
    "视频实际产品SPU",
    "命中方式",
]

MANUAL_SNAPSHOT_HEADERS = [
    "达人ID",
    "统一达人键",
    "复投产品链接",
    "复投产品PID",
    "优先级",
    "下一步动作",
    "负责人",
    "截止日期",
    "备注",
    "来源",
]

RECORD_HEADERS = [
    "达人ID",
    "达人名称",
    "合作日期",
    "产品",
    "视频链接",
    "内容类型",
    "是否出单(Y/N)",
    "订单数",
    "GMV",
    "佣金",
    "是否复投(Y/N)",
    "备注",
]


def resolve_workbook_path() -> Path:
    override = os.environ.get("TIKTOK_SYNC_WORKBOOK_PATH", "").strip()
    if override:
        return Path(override).expanduser()
    return ensure_internal_workbook(DEFAULT_WORKBOOK_PATH)


WORKBOOK_PATH = resolve_workbook_path()
ILLEGAL_EXCEL_CHAR_RE = re.compile(r"[\x00-\x08\x0B-\x0C\x0E-\x1F]")


def normalize_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    return str(value).strip()


def sanitize_excel_text(value: object) -> str:
    text = normalize_text(value)
    if not text:
        return ""
    text = ILLEGAL_EXCEL_CHAR_RE.sub("", text)
    if len(text) > 32767:
        text = text[:32767]
    return text


def sanitize_excel_value(value: object) -> object:
    if value is None:
        return ""
    if isinstance(value, (int, float, datetime)):
        return value
    return sanitize_excel_text(value)


def parse_date(value: object) -> datetime | None:
    text = normalize_text(value)
    if not text:
        return None
    candidates = [text, text[:19], text[:16], text[:10]]
    for candidate in candidates:
        if not candidate:
            continue
        for fmt in (
            "%Y-%m-%d %H:%M:%S",
            "%Y-%m-%d %H:%M",
            "%Y-%m-%d",
            "%Y/%m/%d %H:%M:%S",
            "%Y/%m/%d %H:%M",
            "%Y/%m/%d",
            "%m/%d/%Y %H:%M:%S",
            "%m/%d/%Y %H:%M",
            "%m/%d/%Y",
        ):
            try:
                return datetime.strptime(candidate, fmt)
            except ValueError:
                continue
    return None


def format_date(value: datetime | None) -> str:
    return value.strftime("%Y-%m-%d") if value else ""


def parse_number(value: object) -> float:
    text = normalize_text(value)
    if not text:
        return 0.0
    cleaned = re.sub(r"[^0-9,.-]", "", text)
    if not cleaned:
        return 0.0
    if cleaned.count(",") and not cleaned.count("."):
        cleaned = cleaned.replace(",", "")
    elif cleaned.count(",") and cleaned.count("."):
        cleaned = cleaned.replace(",", "")
    try:
        return float(cleaned)
    except ValueError:
        return 0.0


def load_csv(path: Path) -> list[dict[str, str]]:
    if not path.exists():
        return []
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle))


def resolve_coop_xlsx() -> Path:
    for candidate in COOP_XLSX_CANDIDATES:
        if candidate.exists():
            return candidate
    wechat_root = Path("/Users/apple/Library/Containers/com.tencent.xinWeChat/Data/Documents/xwechat_files")
    if wechat_root.exists():
        matches = sorted(wechat_root.rglob(COOP_XLSX_NAME), reverse=True)
        if matches:
            return matches[0]
    raise FileNotFoundError(f"未找到合作表: {COOP_XLSX_NAME}")


def load_cloud_cooperation_input_rows() -> list[dict[str, object]]:
    workspace = quote(PUBLIC_READ_SYNC_SETTINGS["workspace_id"], safe="")
    try:
        rows = supabase_request(
            f"/tiktok_cooperation_raw?workspace_id=eq.{workspace}&select=cooperation_id,kol_id,platform,cooperation_type,start_at,end_at,is_joint_post,sample_type,shipping_channel,cooperation_attribute,cooperation_fee,prepaid_fee,commission_rate,shipping_address,live_minutes,product_spu_list,created_at_source,updated_at_source,created_by_source,status&order=cooperation_id.asc"
        )
    except (URLError, HTTPError):
        return []
    mapped: list[dict[str, object]] = []
    for row in rows or []:
        mapped.append(
            {
                "合作ID": normalize_text(row.get("cooperation_id")),
                "kolId": normalize_text(row.get("kol_id")),
                "平台": normalize_text(row.get("platform")),
                "合作类型": normalize_text(row.get("cooperation_type")),
                "开始时间": normalize_text(row.get("start_at")),
                "结束时间": normalize_text(row.get("end_at")),
                "是否为Joint Post合作": normalize_text(row.get("is_joint_post")),
                "样品类型": normalize_text(row.get("sample_type")),
                "发货渠道": normalize_text(row.get("shipping_channel")),
                "合作属性": normalize_text(row.get("cooperation_attribute")),
                "合作费用": normalize_text(row.get("cooperation_fee")),
                "预付费用": normalize_text(row.get("prepaid_fee")),
                "佣金比例": normalize_text(row.get("commission_rate")),
                "收获地址": normalize_text(row.get("shipping_address")),
                "直播分钟数": normalize_text(row.get("live_minutes")),
                "合作商品SPU，以 / 分割": normalize_text(row.get("product_spu_list")),
                "创建时间": normalize_text(row.get("created_at_source")),
                "更新时间": normalize_text(row.get("updated_at_source")),
                "创建人": normalize_text(row.get("created_by_source")),
                "状态": normalize_text(row.get("status")),
            }
        )
    return mapped


def load_cooperation_input_rows() -> list[dict[str, object]]:
    cloud_rows = load_cloud_cooperation_input_rows()
    if cloud_rows:
        return cloud_rows
    workbook = load_workbook(resolve_coop_xlsx(), read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    headers = [normalize_text(cell.value) for cell in sheet[1]]
    rows: list[dict[str, object]] = []
    for values in sheet.iter_rows(min_row=2, values_only=True):
        rows.append({headers[idx]: values[idx] for idx in range(min(len(headers), len(values)))})
    workbook.close()
    return rows


def load_cloud_sku_mapping_rows() -> list[dict[str, str]]:
    workspace = quote(PUBLIC_READ_SYNC_SETTINGS["workspace_id"], safe="")
    try:
        rows = supabase_request(
            f"/tiktok_product_sku_cost_raw?workspace_id=eq.{workspace}&select=spu,sku,cost,country_code&order=spu.asc,sku.asc"
        )
    except (URLError, HTTPError):
        return []
    return [
        {
            "spu": normalize_text(row.get("spu")),
            "sku": normalize_text(row.get("sku")).upper(),
            "cost": normalize_text(row.get("cost")),
            "country_code": normalize_text(row.get("country_code")).upper(),
        }
        for row in (rows or [])
        if normalize_text(row.get("spu")) and normalize_text(row.get("sku"))
    ]


def load_pipeline_state() -> dict:
    if not STATE_PATH.exists():
        return {}
    try:
        return json.loads(STATE_PATH.read_text(encoding="utf-8"))
    except json.JSONDecodeError:
        return {}


def write_sheet_rows(sheet, headers: list[str], rows: list[dict[str, object]], start_row: int = 2) -> None:
    if sheet.max_row >= start_row:
        sheet.delete_rows(start_row, sheet.max_row - start_row + 1)
    for row_index, row in enumerate(rows, start=start_row):
        for col_index, header in enumerate(headers, start=1):
            sheet.cell(row=row_index, column=col_index, value=sanitize_excel_value(row.get(header, "")))


def ensure_sheet(workbook, name: str, headers: list[str]):
    if name in workbook.sheetnames:
        sheet = workbook[name]
    else:
        sheet = workbook.create_sheet(name)
    for index, header in enumerate(headers, start=1):
        sheet.cell(row=1, column=index, value=sanitize_excel_text(header))
    return sheet


def supabase_request(path: str) -> list[dict]:
    request = Request(
        f'{PUBLIC_READ_SYNC_SETTINGS["supabase_url"]}/rest/v1{path}',
        headers={
            "apikey": PUBLIC_READ_SYNC_SETTINGS["anon_key"],
            "Authorization": f'Bearer {PUBLIC_READ_SYNC_SETTINGS["anon_key"]}',
        },
    )
    with urlopen(request, timeout=SUPABASE_READ_TIMEOUT_SECONDS) as response:
        return json.loads(response.read().decode("utf-8"))


def load_remote_core_overrides() -> dict[str, dict[str, str]]:
    workspace = quote(PUBLIC_READ_SYNC_SETTINGS["workspace_id"], safe="")
    try:
        rows = supabase_request(f"/creator_sync_overrides?workspace_id=eq.{workspace}&select=kol_id,fields")
    except (URLError, HTTPError, TimeoutError, OSError):
        return {}
    overrides: dict[str, dict[str, str]] = {}
    for row in rows or []:
        raw_fields = row.get("fields") or {}
        fields = {}
        for key, target in CORE_OVERRIDE_MAP.items():
            if key in raw_fields and normalize_text(raw_fields.get(key)):
                fields[target] = normalize_text(raw_fields.get(key))
        if fields:
            overrides[normalize_text(row.get("kol_id"))] = fields
    return overrides


def parse_pid_from_text(value: object) -> str:
    text = normalize_text(value)
    if not text:
        return ""
    if text.isdigit():
        return text
    patterns = [
        r"/product/(\d+)",
        r"[?&](?:pid|product_id)=([0-9]+)",
        r"\b(\d{12,})\b",
    ]
    for pattern in patterns:
        match = re.search(pattern, text)
        if match:
            return match.group(1)
    return ""


def normalize_handle(value: object) -> str:
    text = normalize_text(value).lower().strip()
    text = text.lstrip("@")
    return re.sub(r"[^a-z0-9]", "", text)


def load_creator_pool_tags() -> dict[str, dict[str, str]]:
    if not CREATOR_POOL_JSON.exists():
        return {}
    payload = json.loads(CREATOR_POOL_JSON.read_text(encoding="utf-8"))
    mapping: dict[str, dict[str, str]] = {}
    for creator in payload.get("creators", []):
        creator_id = normalize_text(creator.get("kolId"))
        if not creator_id:
            continue
        mapping[creator_id] = {key: normalize_text(value) for key, value in creator.items()}
    return mapping


def load_level_mapping() -> dict[str, str]:
    if not LEVEL_SHEET_PATH.exists():
        return {}
    workbook = load_workbook(LEVEL_SHEET_PATH, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    headers = [normalize_text(cell.value) for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    level_idx = headers.index("level") if "level" in headers else 0
    creator_idx = 2 if len(headers) >= 3 else 0
    mapping: dict[str, str] = {}
    for row in sheet.iter_rows(min_row=2, values_only=True):
        values = list(row)
        if not values:
            continue
        level = normalize_text(values[level_idx] if level_idx < len(values) else "")
        creator_ref = normalize_text(values[creator_idx] if creator_idx < len(values) else "")
        for key in {normalize_handle(creator_ref), normalize_handle(creator_ref.lstrip("@"))}:
            if key and level:
                mapping[key] = level
    workbook.close()
    return mapping


def load_creator_mapping(workbook) -> tuple[dict[str, str], dict[str, str], dict[str, str]]:
    sheet = workbook["Creator映射"]
    headers = [normalize_text(cell.value) for cell in sheet[1]]
    idx = {header: pos for pos, header in enumerate(headers, start=1)}
    source_to_existing: dict[str, str] = {}
    existing_to_source: dict[str, str] = {}
    existing_to_name: dict[str, str] = {}
    for row in range(2, sheet.max_row + 1):
        source_key = normalize_text(sheet.cell(row, idx.get("统一达人键", 1)).value)
        existing_id = normalize_text(sheet.cell(row, idx.get("现有达人ID", 5)).value)
        existing_name = normalize_text(sheet.cell(row, idx.get("达人名称_现有库", 2)).value)
        if source_key and existing_id:
            source_to_existing[source_key] = existing_id
            existing_to_source[existing_id] = source_key
            existing_to_name[existing_id] = existing_name
    return source_to_existing, existing_to_source, existing_to_name


def load_sku_mapping() -> tuple[dict[str, set[str]], dict[str, str], list[str], list[str]]:
    spu_to_skus: dict[str, set[str]] = defaultdict(set)
    sku_to_spu: dict[str, str] = {}
    all_spus: set[str] = set()
    all_skus: set[str] = set()
    cloud_rows = load_cloud_sku_mapping_rows()
    if cloud_rows:
        raw_rows = [(row.get("spu"), row.get("sku")) for row in cloud_rows]
    else:
        workbook = load_workbook(SKU_XLSX, read_only=True, data_only=True)
        sheet = workbook[workbook.sheetnames[0]]
        raw_rows = list(sheet.iter_rows(min_row=2, values_only=True))
        workbook.close()
    for row in raw_rows:
        spu = normalize_text(row[0])
        sku = normalize_text(row[1]).upper()
        if not spu or not sku:
            continue
        spu_to_skus[spu].add(sku)
        sku_to_spu[sku] = spu
        all_spus.add(spu)
        all_skus.add(sku)
    return spu_to_skus, sku_to_spu, sorted(all_skus, key=len, reverse=True), sorted(all_spus, key=len, reverse=True)


def split_spus(value: object) -> list[str]:
    return [item.strip() for item in normalize_text(value).split("/") if item and item.strip()]


def extract_tokens_from_product_name(name: object, all_skus: list[str], all_spus: list[str], sku_to_spu: dict[str, str]) -> tuple[list[str], list[str]]:
    text = normalize_text(name).upper()
    found_skus = [sku for sku in all_skus if sku and sku in text]
    found_spus = {sku_to_spu[sku] for sku in found_skus if sku in sku_to_spu}
    if not found_spus:
        for spu in all_spus:
            if spu and spu in text:
                found_spus.add(spu)
    return found_skus, sorted(found_spus)


def dedupe_creator_rows() -> list[dict[str, str]]:
    history_rows = load_csv(CREATOR_HISTORY_CSV)
    latest_history_date = max(
        (parse_date(row.get("统计日期")) for row in history_rows if parse_date(row.get("统计日期"))),
        default=None,
    )
    daily_rows = []
    for row in load_csv(CREATOR_CSV):
        stat_date = parse_date(row.get("统计日期"))
        if latest_history_date and stat_date and stat_date <= latest_history_date:
            continue
        daily_rows.append(row)
    seen: set[tuple[str, str, str, str]] = set()
    merged: list[dict[str, str]] = []
    for row in history_rows + daily_rows:
        key = (
            normalize_text(row.get("店铺")),
            normalize_text(row.get("统计日期")),
            normalize_text(row.get("统一达人键")),
            normalize_text(row.get("原文件名")),
        )
        if key in seen:
            continue
        seen.add(key)
        merged.append(row)
    return merged


def aggregate_creator_rows(rows: list[dict[str, str]]) -> dict[str, dict[str, object]]:
    if not rows:
        return {}
    latest_stat = max((parse_date(row.get("统计日期")) for row in rows if parse_date(row.get("统计日期"))), default=datetime.now())
    cutoff_90 = latest_stat - timedelta(days=89)
    grouped: dict[str, dict[str, object]] = defaultdict(
        lambda: {
            "历史总GMV": 0.0,
            "90天GMV": 0.0,
            "累计订单": 0.0,
            "近90天订单": 0.0,
            "累计视频数": 0.0,
            "近90天视频数": 0.0,
            "店铺标签": set(),
            "粉丝量": 0.0,
            "粉丝量统计日期": None,
            "达人名称": "",
            "达人昵称": "",
            "主页链接": "",
            "达人主页标识": "",
            "最近统计日期": "",
        }
    )
    for row in rows:
        key = normalize_text(row.get("统一达人键"))
        if not key:
            continue
        stat_date = parse_date(row.get("统计日期"))
        gmv = parse_number(row.get("Affiliate-attributed GMV"))
        followers = parse_number(row.get("Affiliate followers"))
        orders = parse_number(row.get("Attributed orders"))
        videos = parse_number(row.get("Videos"))
        group = grouped[key]
        group["历史总GMV"] += gmv
        group["累计订单"] += orders
        group["累计视频数"] += videos
        if stat_date and stat_date >= cutoff_90:
            group["90天GMV"] += gmv
            group["近90天订单"] += orders
            group["近90天视频数"] += videos
        group["店铺标签"].add(normalize_text(row.get("店铺")))
        if stat_date:
            date_text = stat_date.strftime("%Y-%m-%d")
            if not normalize_text(group["最近统计日期"]) or date_text > normalize_text(group["最近统计日期"]):
                group["最近统计日期"] = date_text
                group["达人名称"] = normalize_text(row.get("达人名称"))
                group["达人昵称"] = normalize_text(row.get("达人名称"))
                handle = normalize_text(row.get("达人主页标识")) or key
                group["达人主页标识"] = handle
                group["主页链接"] = f"https://www.tiktok.com/@{handle.lstrip('@')}" if handle else ""
        existing_date = group.get("粉丝量统计日期")
        if followers > 0:
            if existing_date is None or (stat_date and stat_date > existing_date):
                group["粉丝量"] = followers
                group["粉丝量统计日期"] = stat_date
            elif stat_date == existing_date and followers > parse_number(group.get("粉丝量")):
                group["粉丝量"] = followers
    return grouped


def preprocess_video_rows(rows: list[dict[str, str]], all_skus: list[str], all_spus: list[str], sku_to_spu: dict[str, str]) -> list[dict[str, object]]:
    enriched: list[dict[str, object]] = []
    for row in rows:
        product_name = normalize_text(row.get("Product name"))
        product_id = parse_pid_from_text(row.get("Product ID"))
        skus, spus = extract_tokens_from_product_name(product_name, all_skus, all_spus, sku_to_spu)
        enriched.append(
            {
                **row,
                "_stat_date": parse_date(row.get("统计日期")),
                "_post_date": parse_date(row.get("Post date")),
                "_gmv": parse_number(row.get("Affiliate video-attributed GMV")),
                "_orders": parse_number(row.get("Video-attributed orders")),
                "_product_pid": product_id,
                "_product_skus": skus,
                "_product_spus": spus,
            }
        )
    return enriched


def aggregate_video_rows(rows: list[dict[str, object]]) -> tuple[dict[str, dict[str, object]], dict[str, list[dict[str, object]]], datetime]:
    latest_reference = max(
        (
            row.get("_post_date") or row.get("_stat_date")
            for row in rows
            if row.get("_post_date") or row.get("_stat_date")
        ),
        default=datetime.now(),
    )
    cutoff_30 = latest_reference - timedelta(days=29)
    cutoff_90 = latest_reference - timedelta(days=89)
    cutoff_1 = latest_reference - timedelta(days=0)
    grouped: dict[str, dict[str, object]] = defaultdict(
        lambda: {
            "店铺标签": set(),
            "近1天GMV": 0.0,
            "近1天订单": 0.0,
            "近1天视频数": 0.0,
            "近30天发布视频GMV": 0.0,
            "近30天订单": 0.0,
            "近30天视频数": 0.0,
            "近90天GMV": 0.0,
            "近90天订单": 0.0,
            "近90天视频数": 0.0,
            "累计GMV": 0.0,
            "累计订单": 0.0,
            "累计视频数": 0.0,
            "最近视频发布时间": "",
            "最近视频发布时间_dt": None,
            "最近出单日期": "",
            "最近统计日期": "",
        }
    )
    videos_by_key: dict[str, list[dict[str, object]]] = defaultdict(list)
    for row in rows:
        key = normalize_text(row.get("统一达人键"))
        if not key:
            continue
        videos_by_key[key].append(row)
        group = grouped[key]
        group["店铺标签"].add(normalize_text(row.get("店铺")))
        stat_date = row.get("_stat_date")
        post_date = row.get("_post_date")
        gmv = row.get("_gmv", 0.0)
        orders = row.get("_orders", 0.0)
        group["累计GMV"] += gmv
        group["累计订单"] += orders
        group["累计视频数"] += 1
        if stat_date and normalize_text(group["最近统计日期"]) < stat_date.strftime("%Y-%m-%d"):
            group["最近统计日期"] = stat_date.strftime("%Y-%m-%d")
        if post_date and (group["最近视频发布时间_dt"] is None or post_date > group["最近视频发布时间_dt"]):
            group["最近视频发布时间_dt"] = post_date
            group["最近视频发布时间"] = post_date.strftime("%Y-%m-%d")
        if post_date and orders > 0 and normalize_text(group["最近出单日期"]) < post_date.strftime("%Y-%m-%d"):
            group["最近出单日期"] = post_date.strftime("%Y-%m-%d")
        if post_date and post_date >= cutoff_1:
            group["近1天GMV"] += gmv
            group["近1天订单"] += orders
            group["近1天视频数"] += 1
        if post_date and post_date >= cutoff_30:
            group["近30天发布视频GMV"] += gmv
            group["近30天订单"] += orders
            group["近30天视频数"] += 1
        if post_date and post_date >= cutoff_90:
            group["近90天GMV"] += gmv
            group["近90天订单"] += orders
            group["近90天视频数"] += 1
    for key, items in videos_by_key.items():
        items.sort(key=lambda item: item.get("_post_date") or datetime.min)
    return grouped, videos_by_key, latest_reference


def map_status(raw_status: str) -> str:
    return "在合作" if raw_status in ACTIVE_COOP_STATUSES else "不在合作"


def build_matchers(
    source_keys: set[str],
    source_to_existing: dict[str, str],
    existing_to_source: dict[str, str],
) -> tuple[dict[str, str], dict[str, str], dict[str, list[str]], dict[str, list[str]]]:
    exact_source = {normalize_text(key): key for key in source_keys}
    exact_existing = {normalize_text(key): value for key, value in existing_to_source.items()}
    clean_to_source: dict[str, list[str]] = defaultdict(list)
    bucket_to_clean: dict[str, list[str]] = defaultdict(list)
    for key in source_keys:
        cleaned = normalize_handle(key)
        if not cleaned:
            continue
        if key not in clean_to_source[cleaned]:
            clean_to_source[cleaned].append(key)
        bucket = cleaned[:2] or cleaned[:1]
        if cleaned not in bucket_to_clean[bucket]:
            bucket_to_clean[bucket].append(cleaned)
    for existing_id, source_key in existing_to_source.items():
        cleaned = normalize_handle(existing_id)
        if cleaned and source_key not in clean_to_source[cleaned]:
            clean_to_source[cleaned].append(source_key)
    return exact_source, exact_existing, clean_to_source, bucket_to_clean


def match_creator_key(
    kol_id: str,
    exact_source: dict[str, str],
    exact_existing: dict[str, str],
    clean_to_source: dict[str, list[str]],
    bucket_to_clean: dict[str, list[str]],
) -> tuple[str, str, float]:
    raw = normalize_text(kol_id)
    if not raw:
        return "", "unmatched", 0.0
    if raw in exact_source:
        return exact_source[raw], "direct_source", 1.0
    if raw in exact_existing:
        return exact_existing[raw], "direct_existing", 1.0
    cleaned = normalize_handle(raw)
    if cleaned in clean_to_source and len(clean_to_source[cleaned]) == 1:
        return clean_to_source[cleaned][0], "normalized_exact", 0.99

    bucket = cleaned[:2] or cleaned[:1]
    candidates = bucket_to_clean.get(bucket, [])
    best_key = ""
    best_score = 0.0
    for candidate_clean in candidates:
        if abs(len(candidate_clean) - len(cleaned)) > 4:
            continue
        score = SequenceMatcher(None, cleaned, candidate_clean).ratio()
        if score > best_score:
            candidate_keys = clean_to_source.get(candidate_clean, [])
            if candidate_keys:
                best_key = candidate_keys[0]
                best_score = score
    if best_key and best_score >= 0.92:
        return best_key, "fuzzy_auto", round(best_score, 4)
    return "", "unmatched", round(best_score, 4)


def load_cooperation_rows(
    source_keys: set[str],
    source_to_existing: dict[str, str],
    existing_to_source: dict[str, str],
    existing_to_name: dict[str, str],
    spu_to_skus: dict[str, set[str]],
) -> tuple[list[dict[str, object]], list[dict[str, object]], list[dict[str, object]]]:
    exact_source, exact_existing, clean_to_source, bucket_to_clean = build_matchers(
        source_keys, source_to_existing, existing_to_source
    )

    mapping_rows: dict[str, dict[str, object]] = {}
    candidate_rows: list[dict[str, object]] = []
    coop_rows: list[dict[str, object]] = []
    for row in load_cooperation_input_rows():
        raw_kol = normalize_text(row.get("kolId"))
        matched_key, method, confidence = match_creator_key(
            raw_kol, exact_source, exact_existing, clean_to_source, bucket_to_clean
        )
        existing_id = source_to_existing.get(matched_key, "")
        mapping_rows.setdefault(
            raw_kol,
            {
                "原始kolId": raw_kol,
                "规范化kolId": normalize_handle(raw_kol),
                "统一达人键": matched_key,
                "现有达人ID": existing_id,
                "达人名称_来源": matched_key,
                "达人名称_现有库": existing_to_name.get(existing_id, ""),
                "匹配方式": method,
                "匹配置信度": confidence,
                "是否自动生效": "Y" if method != "unmatched" else "N",
                "备注": "",
            },
        )
        if method.startswith("fuzzy"):
            candidate_rows.append(
                {
                    "原始kolId": raw_kol,
                    "候选统一达人键": matched_key,
                    "候选现有达人ID": existing_id,
                    "匹配方式": method,
                    "匹配置信度": confidence,
                    "是否已采用": "Y",
                }
            )

        spu_list = split_spus(row.get("合作商品SPU，以 / 分割"))
        sku_list: list[str] = []
        for spu in spu_list:
            sku_list.extend(sorted(spu_to_skus.get(spu, set())))

        coop_rows.append(
            {
                "合作ID": normalize_text(row.get("合作ID")),
                "原始kolId": raw_kol,
                "统一达人键": matched_key,
                "现有达人ID": existing_id,
                "开始时间": parse_date(row.get("开始时间")),
                "结束时间": parse_date(row.get("结束时间")),
                "状态": normalize_text(row.get("状态")),
                "合作费用": parse_number(row.get("合作费用")),
                "预付费用": parse_number(row.get("预付费用")),
                "佣金比例": parse_number(row.get("佣金比例")),
                "合作类型": normalize_text(row.get("合作类型")),
                "合作属性": normalize_text(row.get("合作属性")),
                "合作商品SPU列表": spu_list,
                "合作商品SKU列表": sorted(set(sku_list)),
                "创建时间": parse_date(row.get("创建时间")),
                "更新时间": parse_date(row.get("更新时间")),
            }
        )

    return coop_rows, list(mapping_rows.values()), candidate_rows


def match_video_to_products(
    video: dict[str, object],
    manual_pid: str,
    coop_spus: list[str],
    coop_skus: list[str],
) -> tuple[bool, str]:
    video_pid = normalize_text(video.get("_product_pid"))
    if manual_pid and video_pid and manual_pid == video_pid:
        return True, "manual_pid"
    video_skus = set(video.get("_product_skus", []))
    if coop_skus and video_skus.intersection(coop_skus):
        return True, "sku"
    video_spus = set(video.get("_product_spus", []))
    if coop_spus and video_spus.intersection(coop_spus):
        return True, "spu"
    if manual_pid and manual_pid in normalize_text(video.get("Product name")):
        return True, "manual_pid_name"
    return False, ""


def build_cooperation_metrics(
    coop_rows: list[dict[str, object]],
    videos_by_key: dict[str, list[dict[str, object]]],
    source_to_existing: dict[str, str],
    remote_core_overrides: dict[str, dict[str, str]],
    reference_now: datetime,
) -> tuple[dict[str, dict[str, object]], list[dict[str, object]], list[dict[str, object]]]:
    grouped: dict[str, dict[str, object]] = defaultdict(
        lambda: {
            "最近合作日期": "",
            "当前合作状态": "",
            "近30天合作次数": 0,
            "平均间隔天数": "",
            "距离上次发布天数": "",
            "近30天是否出单(Y/N)": "N",
            "是否进入复投(Y/N)": "N",
            "是否超时未合作": "N",
            "平均交付时间": "",
            "最近合作费用": 0.0,
        }
    )
    product_rows: list[dict[str, object]] = []
    record_rows: list[dict[str, object]] = []
    recent_cutoff = reference_now - timedelta(days=29)

    coops_by_key: dict[str, list[dict[str, object]]] = defaultdict(list)
    for row in coop_rows:
        if normalize_text(row.get("统一达人键")):
            coops_by_key[normalize_text(row.get("统一达人键"))].append(row)

    for creator_key, rows in coops_by_key.items():
        rows.sort(key=lambda item: item.get("开始时间") or datetime.min)
        existing_id = source_to_existing.get(creator_key, "")
        manual_fields = remote_core_overrides.get(existing_id, {})
        manual_link = normalize_text(manual_fields.get("复投产品链接"))
        manual_pid = parse_pid_from_text(manual_fields.get("复投产品PID")) or parse_pid_from_text(manual_link)
        creator_videos = videos_by_key.get(creator_key, [])
        delivery_days: list[int] = []

        for index, coop in enumerate(rows):
            start_time = coop.get("开始时间")
            end_time = coop.get("结束时间")
            next_start = rows[index + 1].get("开始时间") if index + 1 < len(rows) else None
            window_end = next_start or end_time or (start_time + timedelta(days=30) if start_time else None)
            matched_videos = []
            for video in creator_videos:
                post_date = video.get("_post_date")
                if not post_date or not start_time or post_date < start_time:
                    continue
                if window_end and post_date > window_end:
                    continue
                matched, match_mode = match_video_to_products(
                    video, manual_pid, coop.get("合作商品SPU列表", []), coop.get("合作商品SKU列表", [])
                )
                if matched:
                    matched_videos.append((video, match_mode))
            if not matched_videos and start_time:
                fallback = [
                    video
                    for video in creator_videos
                    if video.get("_post_date") and video.get("_post_date") >= start_time and (not window_end or video.get("_post_date") <= window_end)
                ]
                matched_videos = [(video, "post_after_coop") for video in fallback[:1]]

            first_video = matched_videos[0][0] if matched_videos else None
            if start_time and first_video and first_video.get("_post_date"):
                delivery_days.append((first_video["_post_date"] - start_time).days)

            if matched_videos:
                first_video_row, first_mode = matched_videos[0]
                product_rows.append(
                    {
                        "合作ID": coop.get("合作ID", ""),
                        "统一达人键": creator_key,
                        "现有达人ID": existing_id,
                        "开始时间": format_date(start_time),
                        "结束时间": format_date(end_time),
                        "合作状态": coop.get("状态", ""),
                        "合作费用": round(coop.get("合作费用", 0.0), 2),
                        "合作SPU列表": " / ".join(coop.get("合作商品SPU列表", [])),
                        "合作SKU列表": " / ".join(coop.get("合作商品SKU列表", [])),
                        "复投产品链接": manual_link,
                        "复投产品PID": manual_pid,
                        "匹配视频ID": normalize_text(first_video_row.get("Video ID")),
                        "匹配视频链接": normalize_text(first_video_row.get("Video link")),
                        "匹配发布时间": format_date(first_video_row.get("_post_date")),
                        "视频产品名称": normalize_text(first_video_row.get("Product name")),
                        "视频产品PID": normalize_text(first_video_row.get("_product_pid")),
                        "视频实际产品SKU": " / ".join(first_video_row.get("_product_skus", [])),
                        "视频实际产品SPU": " / ".join(first_video_row.get("_product_spus", [])),
                        "命中方式": first_mode,
                    }
                )

            gmv = sum(video.get("_gmv", 0.0) for video, _ in matched_videos)
            orders = sum(video.get("_orders", 0.0) for video, _ in matched_videos)
            video_link = normalize_text(first_video.get("Video link")) if first_video else ""
            record_rows.append(
                {
                    "达人ID": existing_id,
                    "达人名称": creator_key,
                    "合作日期": format_date(start_time),
                    "产品": " / ".join(coop.get("合作商品SPU列表", [])),
                    "视频链接": video_link,
                    "内容类型": coop.get("合作类型", ""),
                    "是否出单(Y/N)": "Y" if orders > 0 or gmv > 0 else "N",
                    "订单数": round(orders, 2),
                    "GMV": round(gmv, 2),
                    "佣金": round(coop.get("合作费用", 0.0), 2),
                    "是否复投(Y/N)": "Y" if coop.get("合作属性") == "Re-coop" else "N",
                    "备注": coop.get("状态", ""),
                }
            )

        latest = rows[-1]
        latest_start = latest.get("开始时间")
        latest_fee = latest.get("合作费用", 0.0)
        post_after_latest = [
            video
            for video in creator_videos
            if video.get("_post_date") and latest_start and video.get("_post_date") > latest_start
        ]
        latest_post = max((video.get("_post_date") for video in post_after_latest), default=None)
        seven_day_deadline = latest_start + timedelta(days=7) if latest_start else None
        matched_within_week = False
        if seven_day_deadline and latest_start:
            for video in creator_videos:
                post_date = video.get("_post_date")
                if not post_date or post_date < latest_start or post_date > seven_day_deadline:
                    continue
                matched, _ = match_video_to_products(
                    video,
                    manual_pid,
                    latest.get("合作商品SPU列表", []),
                    latest.get("合作商品SKU列表", []),
                )
                if matched:
                    matched_within_week = True
                    break

        near30_count = sum(
            1
            for item in rows
            if item.get("开始时间") and item.get("开始时间") >= recent_cutoff
        )
        intervals = []
        for prev, nxt in zip(rows, rows[1:]):
            if prev.get("开始时间") and nxt.get("开始时间"):
                intervals.append((nxt["开始时间"] - prev["开始时间"]).days)

        metrics = grouped[creator_key]
        metrics["最近合作日期"] = format_date(latest_start)
        metrics["当前合作状态"] = map_status(normalize_text(latest.get("状态")))
        metrics["近30天合作次数"] = near30_count
        metrics["平均间隔天数"] = round(sum(intervals) / len(intervals), 1) if intervals else ""
        metrics["距离上次发布天数"] = (reference_now.date() - latest_post.date()).days if latest_post else ""
        metrics["平均交付时间"] = round(sum(delivery_days) / len(delivery_days), 1) if delivery_days else ""
        metrics["最近合作费用"] = latest_fee
        metrics["最近合作SPU列表"] = latest.get("合作商品SPU列表", [])
        metrics["最近合作SKU列表"] = latest.get("合作商品SKU列表", [])
        metrics["复投产品链接"] = manual_link
        metrics["复投产品PID"] = manual_pid

    return grouped, product_rows, record_rows


def refresh_raw_sheet(workbook, sheet_name: str, headers: list[str], rows: list[dict[str, object]], start_row: int = 3) -> None:
    sheet = ensure_sheet(workbook, sheet_name, headers)
    write_sheet_rows(sheet, headers, rows, start_row=start_row)


def update_fact_sheet(
    workbook,
    creator_gmv: dict[str, dict[str, object]],
    video_agg: dict[str, dict[str, object]],
    coop_metrics: dict[str, dict[str, object]],
    remote_core_overrides: dict[str, dict[str, str]],
) -> dict[str, str]:
    sheet = workbook["统一达人事实表"]
    headers = [normalize_text(cell.value) for cell in sheet[1]]
    source_to_existing, creator_ids_to_source, _ = load_creator_mapping(workbook)

    new_headers = [
        "合作表_最近合作日期",
        "合作表_当前合作状态",
        "合作表_近30天合作次数",
        "合作表_平均间隔天数",
        "合作表_距离上次发布天数",
        "合作表_平均交付时间",
        "合作表_最近合作费用",
        "手工_复投产品链接",
        "手工_复投产品PID",
    ]
    for header in new_headers:
        if header not in headers:
            headers.append(header)
            sheet.cell(row=1, column=len(headers), value=sanitize_excel_text(header))

    rows: list[dict[str, object]] = []
    keys = sorted(set(creator_gmv) | set(video_agg) | set(coop_metrics))
    for key in keys:
        creator_stats = creator_gmv.get(key, {})
        video_stats = video_agg.get(key, {})
        coop_stats = coop_metrics.get(key, {})
        existing_id = source_to_existing.get(key, "")
        manual = remote_core_overrides.get(existing_id, {})
        values = {
            "统一达人键": key,
            "现有达人ID": existing_id,
            "达人名称": creator_stats.get("达人名称", ""),
            "达人昵称": creator_stats.get("达人昵称", ""),
            "主页链接": creator_stats.get("主页链接", ""),
            "店铺标签": " / ".join(
                sorted(
                    set(creator_stats.get("店铺标签", set()))
                    | set(video_stats.get("店铺标签", set()))
                )
            ),
            "最近统计日期": creator_stats.get("最近统计日期", "") or video_stats.get("最近统计日期", ""),
            "近1天GMV": round(video_stats.get("近1天GMV", 0.0), 2),
            "近1天订单": round(video_stats.get("近1天订单", 0.0), 2),
            "近1天视频数": int(round(video_stats.get("近1天视频数", 0.0))),
            "近30天GMV": round(video_stats.get("近30天发布视频GMV", 0.0), 2),
            "近30天订单": round(video_stats.get("近30天订单", 0.0), 2),
            "近30天视频数": int(round(video_stats.get("近30天视频数", 0.0))),
            "近90天GMV": round(creator_stats.get("90天GMV", 0.0), 2),
            "近90天订单": round(creator_stats.get("近90天订单", 0.0), 2),
            "近90天视频数": int(round(creator_stats.get("近90天视频数", 0.0))),
            "累计GMV": round(creator_stats.get("历史总GMV", 0.0), 2),
            "累计订单": round(creator_stats.get("累计订单", 0.0), 2),
            "累计视频数": int(round(creator_stats.get("累计视频数", 0.0))),
            "最近视频发布时间": video_stats.get("最近视频发布时间", ""),
            "最近出单日期": video_stats.get("最近出单日期", ""),
            "最近数据来源": "Creator + Videos + Cooperation",
            "合作表_最近合作日期": coop_stats.get("最近合作日期", ""),
            "合作表_当前合作状态": coop_stats.get("当前合作状态", ""),
            "合作表_近30天合作次数": coop_stats.get("近30天合作次数", ""),
            "合作表_平均间隔天数": coop_stats.get("平均间隔天数", ""),
            "合作表_距离上次发布天数": coop_stats.get("距离上次发布天数", ""),
            "合作表_平均交付时间": coop_stats.get("平均交付时间", ""),
            "合作表_最近合作费用": coop_stats.get("最近合作费用", ""),
            "手工_复投产品链接": manual.get("复投产品链接", ""),
            "手工_复投产品PID": parse_pid_from_text(manual.get("复投产品PID")) or parse_pid_from_text(manual.get("复投产品链接")),
            "同步状态": "已同步" if existing_id else "待映射",
        }
        rows.append({header: values.get(header, "") for header in headers})
    write_sheet_rows(sheet, headers, rows, start_row=3)
    return creator_ids_to_source


def update_overview_sheet(
    workbook,
    creator_gmv: dict[str, dict[str, object]],
    video_agg: dict[str, dict[str, object]],
    coop_metrics: dict[str, dict[str, object]],
    existing_to_source: dict[str, str],
    creator_pool_tags: dict[str, dict[str, str]],
    remote_core_overrides: dict[str, dict[str, str]],
) -> None:
    sheet = workbook["达人总览"]
    headers = [normalize_text(cell.value) for cell in sheet[1]]
    level_mapping = load_level_mapping()
    source_to_existing = {source: existing for existing, source in existing_to_source.items() if source}
    now_dt = datetime.now()
    rows: list[dict[str, object]] = []
    keys = sorted(
        set(creator_gmv) | set(video_agg) | set(coop_metrics),
        key=lambda item: (-parse_number(creator_gmv.get(item, {}).get("历史总GMV")), item),
    )
    for source_key in keys:
        creator_id = source_to_existing.get(source_key, source_key)
        creator_stats = creator_gmv.get(source_key, {})
        video_stats = video_agg.get(source_key, {})
        coop_stats = coop_metrics.get(source_key, {})
        seed = creator_pool_tags.get(creator_id, {})
        manual = remote_core_overrides.get(creator_id, {})
        latest_fee = parse_number(coop_stats.get("最近合作费用"))
        near30_video_gmv = parse_number(video_stats.get("近30天发布视频GMV"))
        latest_start = parse_date(coop_stats.get("最近合作日期"))
        timeout = "N"
        if latest_start and near30_video_gmv > latest_fee and latest_fee > 0:
            deadline = latest_start + timedelta(days=7)
            if now_dt > deadline and normalize_text(coop_stats.get("距离上次发布天数")):
                latest_post_date = now_dt - timedelta(days=int(float(normalize_text(coop_stats.get("距离上次发布天数")) or 0)))
                if latest_post_date <= deadline:
                    timeout = "N"
                else:
                    timeout = "Y"
            elif now_dt > deadline and not normalize_text(coop_stats.get("距离上次发布天数")):
                timeout = "Y"
        creator_name = normalize_text(seed.get("达人昵称")) or normalize_text(creator_stats.get("达人名称")) or creator_id
        level = (
            normalize_text(seed.get("合作分层"))
            or level_mapping.get(normalize_handle(creator_id))
            or level_mapping.get(normalize_handle(creator_name))
            or level_mapping.get(normalize_handle(source_key))
        )
        rows.append(
            {
                "达人ID": creator_id,
                "达人名称": creator_name,
                "平台": normalize_text(seed.get("平台")) or "TikTok",
                "粉丝量": int(round(parse_number(creator_stats.get("粉丝量")))) if parse_number(creator_stats.get("粉丝量")) > 0 else "",
                "达人分层(L0/L1/L2/L3)": level,
                "达人类型": normalize_text(seed.get("内容一级标签")),
                "历史总GMV": round(parse_number(creator_stats.get("历史总GMV")), 2),
                "90天GMV": round(parse_number(creator_stats.get("90天GMV")), 2),
                "近30天发布视频gmv": round(near30_video_gmv, 2),
                "最近合作日期": coop_stats.get("最近合作日期", ""),
                "当前合作状态": coop_stats.get("当前合作状态", ""),
                "近30天合作次数": coop_stats.get("近30天合作次数", 0),
                "平均间隔天数": coop_stats.get("平均间隔天数", ""),
                "距离上次发布天数": coop_stats.get("距离上次发布天数", ""),
                "近30天是否出单(Y/N)": "Y" if near30_video_gmv > 0 else "N",
                "是否进入复投(Y/N)": "Y" if near30_video_gmv > latest_fee else "N",
                "是否超时未合作": timeout,
                "平均交付时间": coop_stats.get("平均交付时间", ""),
                "优先级": manual.get("优先级", ""),
                "下一步动作": manual.get("下一步动作", ""),
                "负责人": manual.get("负责人", ""),
                "截止日期": manual.get("截止日期", ""),
                "备注": manual.get("备注", normalize_text(seed.get("备注"))),
            }
        )
    write_sheet_rows(sheet, headers, rows, start_row=4)


def update_focus_sheet(workbook, creator_pool_tags: dict[str, dict[str, str]]) -> None:
    overview = workbook["达人总览"]
    focus = ensure_sheet(
        workbook,
        "L0_L1重点达人池",
        [
            "达人ID",
            "达人名称",
            "平台",
            "粉丝量",
            "达人分层(L0/L1/L2/L3)",
            "达人类型",
            "历史总GMV",
            "近30天GMV",
            "最近合作日期",
            "当前合作状态",
            "近30天合作次数",
            "平均间隔天数",
            "距离上次合作天数",
            "是否出单(Y/N)",
            "是否进入复投(Y/N)",
            "是否超时未合作",
            "优先级",
            "下一步动作",
            "负责人",
            "截止日期",
            "备注",
        ],
    )
    headers = [normalize_text(cell.value) for cell in overview[1]]
    overview_rows = []
    overview_by_id: dict[str, dict[str, object]] = {}
    for row in overview.iter_rows(min_row=4, values_only=True):
        record = {headers[idx]: row[idx] for idx in range(min(len(headers), len(row)))}
        creator_id = normalize_text(record.get("达人ID"))
        if not creator_id:
            continue
        overview_rows.append(record)
        overview_by_id[creator_id] = record

    rows = []
    for creator_id, seed in creator_pool_tags.items():
        record = overview_by_id.get(creator_id, {})
        rows.append(
            {
                "达人ID": creator_id,
                "达人名称": record.get("达人名称", seed.get("达人昵称", creator_id)),
                "平台": record.get("平台", seed.get("平台", "TikTok")),
                "粉丝量": record.get("粉丝量", ""),
                "达人分层(L0/L1/L2/L3)": record.get("达人分层(L0/L1/L2/L3)", ""),
                "达人类型": record.get("达人类型", seed.get("内容一级标签", "")),
                "历史总GMV": record.get("历史总GMV", 0),
                "近30天GMV": record.get("近30天发布视频gmv", 0),
                "最近合作日期": record.get("最近合作日期", ""),
                "当前合作状态": record.get("当前合作状态", seed.get("最近合作状态", "")),
                "近30天合作次数": record.get("近30天合作次数", 0),
                "平均间隔天数": record.get("平均间隔天数", ""),
                "距离上次合作天数": record.get("距离上次发布天数", ""),
                "是否出单(Y/N)": record.get("近30天是否出单(Y/N)", ""),
                "是否进入复投(Y/N)": record.get("是否进入复投(Y/N)", ""),
                "是否超时未合作": record.get("是否超时未合作", ""),
                "优先级": record.get("优先级", ""),
                "下一步动作": record.get("下一步动作", ""),
                "负责人": record.get("负责人", ""),
                "截止日期": record.get("截止日期", ""),
                "备注": record.get("备注", seed.get("备注", "")),
            }
        )
    write_sheet_rows(focus, [normalize_text(cell.value) for cell in focus[1]], rows, start_row=2)


def update_metrics_sheet(workbook) -> None:
    sheet = workbook["运营驾驶舱"]
    overview = workbook["达人总览"]
    pipeline_state = load_pipeline_state()
    stores = pipeline_state.get("stores", {}) if isinstance(pipeline_state, dict) else {}
    last_pipeline_run = pipeline_state.get("last_pipeline_run", {}) if isinstance(pipeline_state, dict) else {}
    headers = [normalize_text(cell.value) for cell in overview[1]]
    rows = []
    for row in overview.iter_rows(min_row=4, values_only=True):
        record = {headers[idx]: row[idx] for idx in range(min(len(headers), len(row)))}
        if normalize_text(record.get("达人ID")):
            rows.append(record)

    l1_l2 = [row for row in rows if normalize_text(row.get("达人分层(L0/L1/L2/L3)")) in {"L1", "L2"}]
    l1_l2_covered = sum(1 for row in l1_l2 if parse_number(row.get("近30天合作次数")) > 0)
    shipped = [row for row in rows if normalize_text(row.get("近30天是否出单(Y/N)")) == "Y"]
    recoop = sum(1 for row in shipped if normalize_text(row.get("是否进入复投(Y/N)")) == "Y")
    avg_freq = round(sum(parse_number(row.get("近30天合作次数")) for row in rows) / len(rows), 2) if rows else 0
    timeout = sum(1 for row in rows if normalize_text(row.get("是否超时未合作")) == "Y")
    product_covered = sum(1 for row in rows if normalize_text(row.get("是否进入复投(Y/N)")) == "Y")

    data_rows = [
        {"指标": "L1/L2合作覆盖率", "数值": f"{(l1_l2_covered / len(l1_l2) * 100):.1f}%" if l1_l2 else "0%", "说明": "L1/L2中近30天有合作的比例", "": "近30天合作次数大于0"},
        {"指标": "出单达人复投率", "数值": f"{(recoop / len(shipped) * 100):.1f}%" if shipped else "0%", "说明": "已出单达人中进入复投的比例", "": ""},
        {"指标": "平均合作频次", "数值": avg_freq, "说明": "达人近30天平均合作次数", "": ""},
        {"指标": "超时达人数量", "数值": timeout, "说明": "超过周期未合作人数", "": ""},
        {"指标": "产品覆盖率", "数值": f"{(product_covered / len(rows) * 100):.1f}%" if rows else "0%", "说明": "当前进入复投的人群覆盖率", "": ""},
        {"指标": "自动化任务状态", "数值": "ACTIVE", "说明": "当前正式日更任务状态", "": "每日 19:00"},
        {"指标": "上次自动化运行", "数值": normalize_text(last_pipeline_run.get("run_at")) or "未记录", "说明": "最近一次流水线运行日期", "": ""},
        {"指标": "上次成功同步", "数值": normalize_text(last_pipeline_run.get("last_success_run_at")) or "未记录", "说明": "最近一次成功完成日更的日期", "": ""},
        {"指标": "当前待同步日期", "数值": normalize_text(last_pipeline_run.get("next_sync_date")) or "未记录", "说明": "四店铺中最早待推进的增量日期", "": ""},
    ]
    last_failure = ""
    for failure in last_pipeline_run.get("failures", []) if isinstance(last_pipeline_run, dict) else []:
        if isinstance(failure, dict):
            last_failure = f"{normalize_text(failure.get('store'))}: {normalize_text(failure.get('message'))}"
            if last_failure.strip(": "):
                break
    data_rows.append({"指标": "最近失败原因", "数值": last_failure or "无", "说明": "如果最近一次没有失败，这里显示无", "": ""})
    success_count = 0
    waiting_count = 0
    failed_count = 0
    for store_key in ("letme", "stypro", "sparco", "icyee"):
        store = stores.get(store_key, {}) if isinstance(stores, dict) else {}
        status = normalize_text(store.get("last_status") or store.get("last_mode")) or "未记录"
        window_status = normalize_text(store.get("window_status"))
        last_daily = normalize_text(store.get("last_daily_date"))
        next_date = normalize_text(store.get("next_increment_date"))
        last_error = normalize_text(store.get("last_error"))
        if status == "success" and last_daily:
            value = f"已同步至 {last_daily}"
            success_count += 1
        elif status == "waiting" and last_daily:
            value = f"已同步至 {last_daily}"
            window_status = "waiting"
            success_count += 1
        elif status == "error":
            value = "运行失败"
            failed_count += 1
        elif status == "skipped":
            value = "已跳过"
        elif status == "bootstrap_required":
            value = "待导入历史基线"
        else:
            value = status
        if window_status == "waiting" and next_date:
            note = f"等待 {next_date} 数据窗口"
            waiting_count += 1
        elif next_date:
            note = f"待同步日期 {next_date}"
        else:
            note = "无待同步日期"
        data_rows.append(
            {
                "指标": f"{store_key.upper()}同步状态",
                "数值": value,
                "说明": note,
                "": last_error,
            }
        )
    data_rows.insert(
        10,
        {
            "指标": "店铺同步结果",
            "数值": f"{success_count} 已同步 / {waiting_count} 等待 / {failed_count} 失败",
            "说明": "当前四店铺同步健康度",
            "": "",
        },
    )
    write_sheet_rows(sheet, [normalize_text(cell.value) for cell in sheet[1]], data_rows, start_row=2)


def append_log(
    workbook,
    status: str,
    message: str,
    video_rows: int,
    creator_rows: int,
    write_meta: dict[str, str] | None = None,
) -> None:
    sheet = workbook["同步日志"]
    headers = [normalize_text(cell.value) for cell in sheet[1]]
    for header in ["写回目标", "写回方式", "校验结果", "桌面发布"]:
        if header not in headers:
            headers.append(header)
            sheet.cell(row=1, column=len(headers), value=sanitize_excel_text(header))
    pipeline_state = load_pipeline_state()
    last_pipeline_run = pipeline_state.get("last_pipeline_run", {})
    stores = pipeline_state.get("stores", {})
    store_summary = []
    for key in ("letme", "stypro", "sparco", "icyee"):
        store_state = stores.get(key, {})
        status_text = normalize_text(store_state.get("last_status") or store_state.get("last_mode"))
        next_date = normalize_text(store_state.get("next_increment_date"))
        error_text = normalize_text(store_state.get("last_error"))
        summary = f"{key}:{status_text or '-'}"
        if next_date:
            summary += f"/next={next_date}"
        if error_text:
            summary += f"/err={error_text}"
        store_summary.append(summary)
    summary_text = " | ".join(store_summary)
    pipeline_text = ""
    if last_pipeline_run:
        pipeline_text = (
            f"run={normalize_text(last_pipeline_run.get('run_at'))}, "
            f"success={normalize_text(last_pipeline_run.get('success'))}, "
            f"skipped={normalize_text(last_pipeline_run.get('skipped'))}, "
            f"failed={normalize_text(last_pipeline_run.get('failed'))}, "
            f"next={normalize_text(last_pipeline_run.get('next_sync_date'))}"
        )
    next_row = sheet.max_row + 1
    sheet.cell(row=next_row, column=1, value=sanitize_excel_text(datetime.now().strftime("%Y-%m-%d %H:%M:%S")))
    run_mode = "自动日更同步" if last_pipeline_run else "手工导入同步"
    sheet.cell(row=next_row, column=2, value=sanitize_excel_text(run_mode))
    sheet.cell(row=next_row, column=3, value="ALL")
    sheet.cell(row=next_row, column=4, value="Videos + Creator + Cooperation")
    sheet.cell(row=next_row, column=5, value="最近可用批次")
    sheet.cell(row=next_row, column=6, value=sanitize_excel_text(f"{VIDEO_CSV.name} / {CREATOR_HISTORY_CSV.name}"))
    sheet.cell(row=next_row, column=7, value="统一达人事实表 / 达人总览 / 合作映射表")
    sheet.cell(row=next_row, column=8, value=video_rows)
    sheet.cell(row=next_row, column=9, value=creator_rows)
    sheet.cell(row=next_row, column=10, value=sanitize_excel_text(status))
    sheet.cell(
        row=next_row,
        column=11,
        value=sanitize_excel_text(" | ".join(part for part in [message, pipeline_text, summary_text] if part)),
    )
    if write_meta:
        sheet.cell(row=next_row, column=headers.index("写回目标") + 1, value=sanitize_excel_text(write_meta.get("write_target", "")))
        sheet.cell(row=next_row, column=headers.index("写回方式") + 1, value=sanitize_excel_text(write_meta.get("write_mode", "")))
        sheet.cell(row=next_row, column=headers.index("校验结果") + 1, value=sanitize_excel_text(write_meta.get("validation_result", "")))
        sheet.cell(row=next_row, column=headers.index("桌面发布") + 1, value=sanitize_excel_text(write_meta.get("desktop_publish", "")))


def finalize_last_log_write_meta(path: Path, write_meta: dict[str, str]) -> None:
    workbook = load_workbook(path)
    sheet = workbook["同步日志"]
    headers = [normalize_text(cell.value) for cell in sheet[1]]
    last_row = sheet.max_row
    if last_row < 3:
        workbook.close()
        return
    mapping = {
        "写回目标": write_meta.get("write_target", ""),
        "写回方式": write_meta.get("write_mode", ""),
        "校验结果": write_meta.get("validation_result", ""),
        "桌面发布": write_meta.get("desktop_publish", ""),
    }
    for header, value in mapping.items():
        if header not in headers:
            headers.append(header)
            sheet.cell(row=1, column=len(headers), value=sanitize_excel_text(header))
        sheet.cell(row=last_row, column=headers.index(header) + 1, value=sanitize_excel_text(value))
    result = save_workbook_safely(workbook, publish_desktop=True)
    workbook.close()
    print(f"finalized sync log: {result}")


def parse_pipeline_count(payload: dict, key: str) -> int:
    try:
        return int(payload.get(key) or 0)
    except (TypeError, ValueError):
        return 0


def derive_log_status(last_pipeline_run: dict) -> str:
    success = parse_pipeline_count(last_pipeline_run, "success")
    waiting = parse_pipeline_count(last_pipeline_run, "waiting")
    skipped = parse_pipeline_count(last_pipeline_run, "skipped")
    failed = parse_pipeline_count(last_pipeline_run, "failed")
    if failed and not success and not skipped and not waiting:
        return "失败"
    if failed:
        return "部分失败"
    if skipped and success:
        return "部分跳过"
    if skipped:
        return "已跳过"
    if waiting and not success:
        return "等待窗口"
    return "成功"


def derive_log_message(base_message: str, last_pipeline_run: dict) -> str:
    success = parse_pipeline_count(last_pipeline_run, "success")
    skipped = parse_pipeline_count(last_pipeline_run, "skipped")
    failed = parse_pipeline_count(last_pipeline_run, "failed")
    if failed:
        return f"本次导出存在失败，沿用当前标准明细 | {base_message}"
    if skipped and success:
        return f"本次部分店铺跳过，已用成功店铺结果刷新标准明细 | {base_message}"
    if skipped:
        return f"本次未新增导入，沿用当前标准明细 | {base_message}"
    return base_message


def main() -> None:
    video_rows_raw = load_csv(VIDEO_CSV)
    creator_rows_raw = dedupe_creator_rows()
    creator_pool_tags = load_creator_pool_tags()
    remote_core_overrides = load_remote_core_overrides()
    spu_to_skus, sku_to_spu, all_skus, all_spus = load_sku_mapping()

    video_rows = preprocess_video_rows(video_rows_raw, all_skus, all_spus, sku_to_spu)
    video_agg, videos_by_key, latest_video_ref = aggregate_video_rows(video_rows)
    creator_gmv = aggregate_creator_rows(creator_rows_raw)

    if SOURCE_WORKBOOK.exists():
        workbook = load_workbook(SOURCE_WORKBOOK)
    else:
        workbook = Workbook()
        if workbook.active:
            workbook.active.title = "达人总览"
    build_notes_sheet(workbook)
    build_sync_sheets(workbook)
    seed_creator_mapping(workbook)
    source_to_existing, existing_to_source, existing_to_name = load_creator_mapping(workbook)
    coop_rows, mapping_rows, candidate_rows = load_cooperation_rows(
        set(creator_gmv) | set(video_agg),
        source_to_existing,
        existing_to_source,
        existing_to_name,
        spu_to_skus,
    )
    coop_metrics, product_rows, record_rows = build_cooperation_metrics(
        coop_rows,
        videos_by_key,
        source_to_existing,
        remote_core_overrides,
        latest_video_ref,
    )

    refresh_raw_sheet(workbook, "原始视频表现", RAW_VIDEO_HEADERS, video_rows_raw)
    refresh_raw_sheet(workbook, "原始达人表现", RAW_CREATOR_HEADERS, creator_rows_raw)
    refresh_raw_sheet(workbook, "合作映射表", COOP_MAPPING_HEADERS, mapping_rows, start_row=2)
    refresh_raw_sheet(workbook, "自动匹配候选", AUTO_CANDIDATE_HEADERS, candidate_rows, start_row=2)
    refresh_raw_sheet(workbook, "产品映射结果", PRODUCT_RESULT_HEADERS, product_rows, start_row=2)

    manual_snapshot_rows = []
    for creator_id, fields in sorted(remote_core_overrides.items()):
        source_key = existing_to_source.get(creator_id, "")
        manual_snapshot_rows.append(
            {
                "达人ID": creator_id,
                "统一达人键": source_key,
                "复投产品链接": fields.get("复投产品链接", ""),
                "复投产品PID": parse_pid_from_text(fields.get("复投产品PID")) or parse_pid_from_text(fields.get("复投产品链接")),
                "优先级": fields.get("优先级", ""),
                "下一步动作": fields.get("下一步动作", ""),
                "负责人": fields.get("负责人", ""),
                "截止日期": fields.get("截止日期", ""),
                "备注": fields.get("备注", ""),
                "来源": "creator_sync_overrides",
            }
        )
    refresh_raw_sheet(workbook, "人工维护快照", MANUAL_SNAPSHOT_HEADERS, manual_snapshot_rows, start_row=2)

    existing_to_source = update_fact_sheet(workbook, creator_gmv, video_agg, coop_metrics, remote_core_overrides)
    update_overview_sheet(
        workbook,
        creator_gmv,
        video_agg,
        coop_metrics,
        existing_to_source,
        creator_pool_tags,
        remote_core_overrides,
    )
    refresh_raw_sheet(workbook, "合作记录明细", RECORD_HEADERS, record_rows, start_row=2)
    update_focus_sheet(workbook, creator_pool_tags)
    update_metrics_sheet(workbook)
    pipeline_state = load_pipeline_state()
    last_pipeline_run = pipeline_state.get("last_pipeline_run", {}) if isinstance(pipeline_state, dict) else {}
    base_message = f"videos={len(video_rows_raw)}, creators={len(creator_rows_raw)}, coop={len(coop_rows)}, product={len(product_rows)}"

    append_log(
        workbook,
        status=derive_log_status(last_pipeline_run),
        message=derive_log_message(base_message, last_pipeline_run),
        video_rows=len(video_rows_raw),
        creator_rows=len(creator_rows_raw),
        write_meta={
            "write_target": f"{INTERNAL_WORKBOOK_PATH} -> {DESKTOP_WORKBOOK_PATH}",
            "write_mode": "staging-validate-atomic-replace",
            "validation_result": "待校验",
            "desktop_publish": "待发布",
        },
    )
    result = save_workbook_safely(workbook, publish_desktop=True)
    finalize_last_log_write_meta(INTERNAL_WORKBOOK_PATH, result)
    print(f"synced {WORKBOOK_PATH}")
    print(result)


if __name__ == "__main__":
    main()
