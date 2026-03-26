from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook, load_workbook


LEVEL_WORKBOOK = Path("/Users/apple/Downloads/达人等级底表/达人等级底表.xlsx")
MAIN_WORKBOOK = Path("/Users/apple/Desktop/达人多次合作监控看板_同步版.xlsx")

LEVEL_SHEET_NAME = "达人等级映射"


def normalize_text(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def to_int(value: object) -> int:
    text = normalize_text(value)
    if not text:
        return 0
    try:
        return int(float(text))
    except ValueError:
        return 0


def timeout_threshold(level: str) -> int:
    return {
        "L0": 14,
        "L1": 21,
        "L2": 30,
        "L3": 45,
    }.get(level, 30)


def derive_timeout(level: str, current_status: str, repeat_flag: str, days_since_last: int) -> str:
    if current_status == "在合作":
        return "N"
    if repeat_flag != "Y":
        return "N"
    if not days_since_last:
        return "N"
    return "Y" if days_since_last > timeout_threshold(level) else "N"


def derive_priority(level: str, current_status: str, timeout_flag: str, recent_30_count: int) -> str:
    if timeout_flag == "Y":
        return "高"
    if level in {"L0", "L1"} and (current_status == "在合作" or recent_30_count > 0):
        return "高"
    if current_status == "在合作" or recent_30_count > 0:
        return "中"
    return "低"


def derive_next_action(
    level: str,
    current_status: str,
    timeout_flag: str,
    repeat_flag: str,
    recent_30_count: int,
) -> str:
    if current_status == "在合作":
        return "跟进交付和发布时间"
    if timeout_flag == "Y":
        return "优先重启复投沟通"
    if level == "L0":
        return "安排重点达人复盘"
    if repeat_flag == "Y":
        return "推进下一轮复投"
    if recent_30_count > 0:
        return "补充选品并发起合作"
    return "观察内容并建立联系"


def load_level_rows() -> list[dict[str, str]]:
    workbook = load_workbook(LEVEL_WORKBOOK, read_only=True, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    rows: list[dict[str, str]] = []
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        creator_id = normalize_text(row[2] if len(row) > 2 else "")
        if not creator_id:
            continue
        rows.append(
            {
                "level": normalize_text(row[0] if len(row) > 0 else ""),
                "brand": normalize_text(row[1] if len(row) > 1 else ""),
                "creator_id": creator_id,
                "contact": normalize_text(row[3] if len(row) > 3 else ""),
            }
        )
    return rows


def main() -> None:
    level_rows = load_level_rows()
    workbook = load_workbook(MAIN_WORKBOOK)
    overview = workbook["达人总览"]
    headers = [normalize_text(cell.value) for cell in overview[1]]
    col = {header: index + 1 for index, header in enumerate(headers)}

    overview_index: dict[str, int] = {}
    for row_number in range(4, overview.max_row + 1):
        creator_id = normalize_text(overview.cell(row=row_number, column=col["达人ID"]).value)
        if creator_id:
            overview_index[creator_id] = row_number

    mapping_rows: list[list[object]] = [
        ["达人ID", "品牌", "等级底表分层", "contact", "原分层", "新分层", "是否命中"],
    ]

    matched = 0
    for level_row in level_rows:
        creator_id = level_row["creator_id"]
        row_number = overview_index.get(creator_id)
        if row_number is None:
            mapping_rows.append(
                [creator_id, level_row["brand"], level_row["level"], level_row["contact"], "", "", "未命中"]
            )
            continue

        old_level = normalize_text(overview.cell(row=row_number, column=col["达人分层(L0/L1/L2/L3)"]).value)
        new_level = level_row["level"] or old_level or "L3"
        overview.cell(row=row_number, column=col["达人分层(L0/L1/L2/L3)"], value=new_level)

        current_status = normalize_text(overview.cell(row=row_number, column=col["当前合作状态"]).value)
        repeat_flag = normalize_text(overview.cell(row=row_number, column=col["是否进入复投(Y/N)"]).value)
        recent_30_count = to_int(overview.cell(row=row_number, column=col["近30天合作次数"]).value)
        days_since_last = to_int(overview.cell(row=row_number, column=col["距离上次发布天数"]).value)

        timeout_flag = derive_timeout(new_level, current_status, repeat_flag, days_since_last)
        priority = derive_priority(new_level, current_status, timeout_flag, recent_30_count)
        next_action = derive_next_action(new_level, current_status, timeout_flag, repeat_flag, recent_30_count)

        overview.cell(row=row_number, column=col["是否超时未合作"], value=timeout_flag)
        overview.cell(row=row_number, column=col["优先级"], value=priority)
        overview.cell(row=row_number, column=col["下一步动作"], value=next_action)

        mapping_rows.append(
            [creator_id, level_row["brand"], level_row["level"], level_row["contact"], old_level, new_level, "命中"]
        )
        matched += 1

    if LEVEL_SHEET_NAME in workbook.sheetnames:
        del workbook[LEVEL_SHEET_NAME]
    mapping_sheet = workbook.create_sheet(LEVEL_SHEET_NAME)
    for row in mapping_rows:
        mapping_sheet.append(row)

    workbook.save(MAIN_WORKBOOK)
    print(f"updated {MAIN_WORKBOOK}")
    print(f"matched={matched}")
    print(f"unmatched={len(level_rows) - matched}")


if __name__ == "__main__":
    main()
