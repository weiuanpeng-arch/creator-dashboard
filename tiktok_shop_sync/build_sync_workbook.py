from __future__ import annotations

from pathlib import Path
import json

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

from sync_schema import SYNC_SHEETS, workbook_notes
from workbook_io import DESKTOP_WORKBOOK_PATH, INTERNAL_WORKBOOK_PATH, save_workbook_safely


BASE_DIR = Path(__file__).resolve().parent
SOURCE_WORKBOOK = Path("/Users/apple/Desktop/达人多次合作监控看板(1).xlsx")
OUTPUT_WORKBOOK = INTERNAL_WORKBOOK_PATH
CREATOR_POOL_JSON = Path("/Users/apple/Documents/Playground/creator_dashboard/data/creator_pool.json")
LEGACY_DASHBOARD_JSON = Path("/Users/apple/Documents/Playground/creator_dashboard/data/core_creator_dashboard.json")

BRAND_TO_STORE = {
    "LetMe": "Letme Home Living",
    "Stypro": "STYPRO.ID",
    "Sparco": "spar.co jewelry",
    "ICYEE": "Icyee Indonesia",
}


def ensure_sheet(workbook: Workbook, name: str):
    if name in workbook.sheetnames:
        ws = workbook[name]
        workbook.remove(ws)
    return workbook.create_sheet(title=name)


def write_header_row(ws, headers: list[str]) -> None:
    for col_index, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_index, value=header)
        cell.font = Font(bold=True)
    ws.freeze_panes = "A2"


def build_notes_sheet(workbook: Workbook) -> None:
    ws = ensure_sheet(workbook, "同步说明")
    ws["A1"] = "字段"
    ws["B1"] = "说明"
    ws["A1"].font = Font(bold=True)
    ws["B1"].font = Font(bold=True)
    for row_index, note in enumerate(workbook_notes(), start=2):
        ws.cell(row=row_index, column=1, value=f"说明{row_index - 1}")
        ws.cell(row=row_index, column=2, value=note)
    ws.freeze_panes = "A2"


def build_sync_sheets(workbook: Workbook) -> None:
    for schema in SYNC_SHEETS:
        ws = ensure_sheet(workbook, schema.name)
        write_header_row(ws, schema.headers)
        ws.cell(row=2, column=1, value="说明")
        ws.cell(row=2, column=2, value=schema.description)


def normalize_text(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def normalize_creator_key(creator: dict[str, object]) -> str:
    profile = normalize_text(creator.get("主页链接"))
    if "/@" in profile:
        return profile.rsplit("/@", 1)[-1].split("?", 1)[0].lower()
    kol_id = normalize_text(creator.get("kolId"))
    return kol_id.lower()


def seed_creator_mapping(workbook: Workbook) -> None:
    if not CREATOR_POOL_JSON.exists():
        return
    payload = json.loads(CREATOR_POOL_JSON.read_text(encoding="utf-8"))
    creators = payload.get("creators", [])

    mapping_ws = workbook["Creator映射"]
    fact_ws = workbook["统一达人事实表"]

    mapping_row = 3
    fact_row = 3
    for creator in creators:
        creator_key = normalize_creator_key(creator)
        brand = normalize_text(creator.get("适配品牌"))
        store_tag = BRAND_TO_STORE.get(brand, "")

        mapping_values = [
            creator_key,
            normalize_text(creator.get("达人昵称")) or normalize_text(creator.get("kolId")),
            normalize_text(creator.get("达人昵称")),
            normalize_text(creator.get("主页链接")),
            normalize_text(creator.get("kolId")),
            normalize_text(creator.get("达人昵称")),
            creator_key,
            "现有达人库预置",
            "1.00",
            store_tag,
            brand,
            normalize_text(creator.get("最近跟进人")),
            normalize_text(creator.get("备注")),
        ]
        for col_index, value in enumerate(mapping_values, start=1):
            mapping_ws.cell(row=mapping_row, column=col_index, value=value)
        mapping_row += 1

        fact_values = [
            creator_key,
            normalize_text(creator.get("kolId")),
            normalize_text(creator.get("达人昵称")) or normalize_text(creator.get("kolId")),
            normalize_text(creator.get("达人昵称")),
            normalize_text(creator.get("主页链接")),
            store_tag,
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "现有达人库",
            "待同步",
            normalize_text(creator.get("备注")),
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
        ]
        for col_index, value in enumerate(fact_values, start=1):
            fact_ws.cell(row=fact_row, column=col_index, value=value)
        fact_row += 1


def seed_dashboard_sheets(workbook: Workbook) -> None:
    if not LEGACY_DASHBOARD_JSON.exists():
        return
    payload = json.loads(LEGACY_DASHBOARD_JSON.read_text(encoding="utf-8"))

    overview_ws = workbook["达人总览"]
    focus_ws = workbook["L0_L1重点达人池"]
    records_ws = workbook["合作记录明细"]
    metrics_ws = workbook["运营驾驶舱"]

    overview_headers = [cell.value for cell in overview_ws[1]]
    focus_headers = [cell.value for cell in focus_ws[1]]
    record_headers = [cell.value for cell in records_ws[1]]
    metric_headers = [cell.value for cell in metrics_ws[1]]

    def map_overview_row(row: dict[str, object]) -> list[object]:
        mapped = {
            "达人ID": row.get("达人ID", ""),
            "达人名称": row.get("达人名称", ""),
            "平台": row.get("平台", ""),
            "粉丝量": row.get("粉丝量", ""),
            "达人分层(L0/L1/L2/L3)": row.get("达人分层(L0/L1/L2/L3)", ""),
            "达人类型": row.get("达人类型", ""),
            "历史总GMV": row.get("历史总GMV", ""),
            "90天GMV": "",
            "近30天发布视频gmv": "",
            "最近合作日期": row.get("最近合作日期", ""),
            "当前合作状态": row.get("当前合作状态", ""),
            "近30天合作次数": row.get("近30天合作次数", ""),
            "平均间隔天数": row.get("平均间隔天数", ""),
            "距离上次发布天数": row.get("距离上次发布天数", ""),
            "近30天是否出单(Y/N)": row.get("是否出单(Y/N)", ""),
            "是否进入复投(Y/N)": row.get("是否进入复投(Y/N)", ""),
            "是否超时未合作": row.get("是否超时未合作", ""),
            "平均交付时间": row.get("平均交付时间", ""),
            "优先级": row.get("优先级", ""),
            "下一步动作": row.get("下一步动作", ""),
            "负责人": row.get("负责人", ""),
            "截止日期": row.get("截止日期", ""),
            "备注": row.get("备注", ""),
        }
        return [mapped.get(header, "") for header in overview_headers]

    def map_focus_row(row: dict[str, object]) -> list[object]:
        return [row.get(header, "") for header in focus_headers]

    def map_record_row(row: dict[str, object]) -> list[object]:
        return [row.get(header, "") for header in record_headers]

    def map_metric_row(row: dict[str, object]) -> list[object]:
        return [row.get(header, "") for header in metric_headers]

    for row_index, row in enumerate(payload.get("overview", []), start=4):
        for col_index, value in enumerate(map_overview_row(row), start=1):
            overview_ws.cell(row=row_index, column=col_index, value=value)

    for row_index, row in enumerate(payload.get("focusPool", []), start=2):
        for col_index, value in enumerate(map_focus_row(row), start=1):
            focus_ws.cell(row=row_index, column=col_index, value=value)

    for row_index, row in enumerate(payload.get("records", []), start=2):
        for col_index, value in enumerate(map_record_row(row), start=1):
            records_ws.cell(row=row_index, column=col_index, value=value)

    for row_index, row in enumerate(payload.get("metrics", []), start=2):
        for col_index, value in enumerate(map_metric_row(row), start=1):
            metrics_ws.cell(row=row_index, column=col_index, value=value)


def main() -> None:
    if SOURCE_WORKBOOK.exists():
        workbook = load_workbook(SOURCE_WORKBOOK)
    else:
        workbook = Workbook()
        if workbook.active:
            workbook.active.title = "达人总览"
    build_notes_sheet(workbook)
    build_sync_sheets(workbook)
    seed_dashboard_sheets(workbook)
    seed_creator_mapping(workbook)
    result = save_workbook_safely(workbook, publish_desktop=True)
    print(f"saved internal workbook {OUTPUT_WORKBOOK}")
    print(f"published desktop workbook {DESKTOP_WORKBOOK_PATH}")
    print(result)


if __name__ == "__main__":
    main()
