from __future__ import annotations

import os
import shutil
import tempfile
from pathlib import Path

from openpyxl import load_workbook


DESKTOP_WORKBOOK_PATH = Path("/Users/apple/Desktop/达人多次合作监控看板_同步版.xlsx")
INTERNAL_WORKBOOK_PATH = Path("/Users/apple/Documents/Playground/tiktok_shop_sync/data/达人多次合作监控看板_同步版.xlsx")
STAGING_DIR = INTERNAL_WORKBOOK_PATH.parent / "_staging"

REQUIRED_SHEET_HEADERS = {
    "达人总览": ["达人ID", "达人名称"],
    "运营驾驶舱": ["指标", "数值"],
    "同步日志": ["运行时间", "运行模式"],
    "合作映射表": ["原始kolId", "统一达人键"],
    "产品映射结果": ["合作ID", "统一达人键"],
}


def normalize_text(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def validate_workbook(path: Path) -> dict[str, object]:
    workbook = load_workbook(path, read_only=True, data_only=False)
    try:
        sheetnames = set(workbook.sheetnames)
        missing = [name for name in REQUIRED_SHEET_HEADERS if name not in sheetnames]
        if missing:
            raise ValueError(f"missing sheets: {', '.join(missing)}")
        for name, required_headers in REQUIRED_SHEET_HEADERS.items():
            worksheet = workbook[name]
            headers = [normalize_text(cell.value) for cell in worksheet[1]]
            for required in required_headers:
                if required not in headers:
                    raise ValueError(f"{name} missing header: {required}")
        return {
            "valid": True,
            "sheet_count": len(workbook.sheetnames),
            "checked_sheets": list(REQUIRED_SHEET_HEADERS.keys()),
        }
    finally:
        workbook.close()


def atomic_replace_file(source: Path, destination: Path) -> None:
    destination.parent.mkdir(parents=True, exist_ok=True)
    tmp_target = destination.parent / f".{destination.name}.tmp"
    if tmp_target.exists():
        tmp_target.unlink()
    shutil.copy2(source, tmp_target)
    os.replace(tmp_target, destination)


def ensure_internal_workbook(seed_path: Path | None = None) -> Path:
    INTERNAL_WORKBOOK_PATH.parent.mkdir(parents=True, exist_ok=True)
    if INTERNAL_WORKBOOK_PATH.exists():
        validate_workbook(INTERNAL_WORKBOOK_PATH)
        return INTERNAL_WORKBOOK_PATH

    candidate = seed_path or DESKTOP_WORKBOOK_PATH
    if not candidate.exists():
        raise FileNotFoundError(f"missing workbook seed: {candidate}")
    validate_workbook(candidate)
    atomic_replace_file(candidate, INTERNAL_WORKBOOK_PATH)
    return INTERNAL_WORKBOOK_PATH


def save_workbook_safely(workbook, *, publish_desktop: bool = True) -> dict[str, str]:
    INTERNAL_WORKBOOK_PATH.parent.mkdir(parents=True, exist_ok=True)
    STAGING_DIR.mkdir(parents=True, exist_ok=True)
    fd, tmp_name = tempfile.mkstemp(prefix="sync-workbook-", suffix=".xlsx", dir=STAGING_DIR)
    os.close(fd)
    temp_path = Path(tmp_name)

    desktop_result = "未发布"
    validation_result = "未校验"
    try:
        workbook.save(temp_path)
        validation = validate_workbook(temp_path)
        validation_result = f"通过({validation['sheet_count']} sheets)"
        atomic_replace_file(temp_path, INTERNAL_WORKBOOK_PATH)
        if publish_desktop:
            atomic_replace_file(INTERNAL_WORKBOOK_PATH, DESKTOP_WORKBOOK_PATH)
            validate_workbook(DESKTOP_WORKBOOK_PATH)
            desktop_result = "成功"
        return {
            "internal_path": str(INTERNAL_WORKBOOK_PATH),
            "desktop_path": str(DESKTOP_WORKBOOK_PATH),
            "write_target": f"{INTERNAL_WORKBOOK_PATH} -> {DESKTOP_WORKBOOK_PATH}",
            "write_mode": "staging-validate-atomic-replace",
            "validation_result": validation_result,
            "desktop_publish": desktop_result,
        }
    finally:
        if temp_path.exists():
            temp_path.unlink()
